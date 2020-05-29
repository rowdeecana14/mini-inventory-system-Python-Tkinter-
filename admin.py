from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from datetime import date
from time import strftime
from tkcalendar import Calendar, DateEntry
from winreg import *
from pandas import DataFrame
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from controller import Controller

"""
#====================================================================================================================#
#-----------------------            MAIN CLASS                                                -----------------------#
#====================================================================================================================#
"""
class Main:
    def __init__(self):
        pass

    def show(self):
        ROOT.withdraw()

        global FRAME_MAIN, MAIN_WIDTH, MAIN_HEIGHT
        MAIN_WIDTH = 1000
        MAIN_HEIGHT = 700
        FRAME_MAIN = Toplevel(ROOT)
        config = window_config(FRAME_MAIN, MAIN_WIDTH, MAIN_HEIGHT)

        # ---- HEADER FRAME ----#
        self.top_nav()

        # ---- BOTTOM FRAME ----#
        self.footer()

        # ---- SIDE FRAME ----#
        self.side_nav()

        # ---- MAIN FRAME ----#
        self.main_body()
        FRAME_MAIN.geometry('%dx%d+%d+%d' % (config['WIDTH'], config['HEIGHT'], config['x'], config['y']))
        FRAME_MAIN.wm_protocol("WM_DELETE_WINDOW", self.close)

    def top_nav(self):
        FRAME_TOP = Frame(FRAME_MAIN, bg="blue")
        FRAME_TOP.place(relx=0, rely=0, relwidth=1, relheight=0.09)

        FRAME_INNER = Frame(FRAME_MAIN, bg="#5cb7d8")
        FRAME_INNER.place(relx=0, rely=0, relwidth=1, relheight=0.08)

        global ICON_SYSTEM
        ICON_SYSTEM = PhotoImage(file="icon/system_logo.png")
        LABEL_ICON_SYSTEM = Label(FRAME_INNER, image=ICON_SYSTEM, bg="#5cb7d8")
        LABEL_ICON_SYSTEM.place(x=20, y=0)

        LABEL_SYSTEM_TITLE = Label(FRAME_INNER, text="STORE INVENTORY SYSTEM", bg="#5cb7d8", font=("algerian", 20))
        LABEL_SYSTEM_TITLE.grid(row=0, column=0, padx=85, ipady=11)

        global ICON_CALENDAR
        ICON_CALENDAR = PhotoImage(file="icon/clock.png")
        LABEL_ICON_CALENDAR = Label(FRAME_INNER, image=ICON_CALENDAR, bg="#5cb7d8")
        LABEL_ICON_CALENDAR.place(relx=0.8, rely=0, relwidth=0.05, relheight=1)

        global LABEL_TIME
        LABEL_TIME = Label(FRAME_INNER, bg="#5cb7d8", font=("arial", 16, "bold"))
        LABEL_TIME.place(relx=0.84, rely=0, relwidth=0.13, relheight=1)
        self.time()

    def side_nav(self):
        global FRAME_SIDE
        FRAME_SIDE = Frame(FRAME_MAIN, bg="#928c8c")

        # ---- SIDE USER LOG0 FRAME ----#
        global FRAME_USER
        FRAME_USER = Frame(FRAME_MAIN, bg="white")

        # ---- SIDE USER LOG0 IMAGE ----#
        global USER_LOGO
        USER_LOGO = PhotoImage(file="icon/user.png")

        global USER_LOGO_LABEL
        USER_LOGO_LABEL = Label(FRAME_MAIN, image=USER_LOGO, bg="black")

        global USER_POSITION_LABEL
        USER_POSITION_LABEL = Label(FRAME_MAIN, text="Administrator", font=("algerian", 11, "bold"), bg="#e68383", fg="black")

        # ---- FRAME MENU ----#
        global FRAME_MENU
        FRAME_MENU = Frame(FRAME_MAIN, bg="white")

        global LABEL_MENU
        LABEL_MENU = Label(FRAME_MENU, text="Menu:", font=("algerian", 12), bg="black", fg="white")

        # ---- FRAME BUTTONS1 ----#
        global FRAME_BUTTON_HOME, BUTTON_HOME
        FRAME_BUTTON_HOME = Frame(FRAME_MAIN, bg="#ddd")
        BUTTON_HOME = Button(FRAME_BUTTON_HOME, text="HOME", bg="blue", fg="white", font=("arial", 10, "bold"), height=1, command=self.home)

        # ---- FRAME BUTTONS2 ----#
        global FRAME_BUTTON_PRODUCT, BUTTON_PRODUCT
        FRAME_BUTTON_PRODUCT = Frame(FRAME_MAIN, bg="#ddd")
        BUTTON_PRODUCT = Button(FRAME_BUTTON_PRODUCT, text="ADD STOCK", bg="#1d1dab", fg="white", font=("arial", 10, "bold"), command=self.product)

        # ---- FRAME BUTTONS3 ----#
        global FRAME_BUTTON_INVENTORY, BUTTON_INVENTORY
        FRAME_BUTTON_INVENTORY = Frame(FRAME_MAIN, bg="#ddd")
        BUTTON_INVENTORY = Button(FRAME_BUTTON_INVENTORY, text="INVENTORY", bg="#1d1dab", fg="white", font=("arial", 10, "bold"), command=self.inventory)

        # ---- FRAME BUTTONS4 ----#
        global FRAME_BUTTON_INVOICE, BUTTON_INVOICE
        FRAME_BUTTON_INVOICE = Frame(FRAME_MAIN, bg="#ddd")
        BUTTON_INVOICE = Button(FRAME_BUTTON_INVOICE, text="RECORDS", bg="#1d1dab", fg="white", font=("arial", 10, "bold"), command=self.invoice)

        # ---- FRAME BUTTONS5 ----#
        global FRAME_BUTTON_USER, BUTTON_USER
        FRAME_BUTTON_USER = Frame(FRAME_MAIN, bg="#ddd")
        BUTTON_USER = Button(FRAME_BUTTON_USER, text="USERS", bg="#1d1dab", fg="white", font=("arial", 10, "bold"), command=self.user)

        # ---- FRAME BUTTONS6 ----#
        global FRAME_BUTTON_LOGOUT, BUTTON_LOGOUT
        FRAME_BUTTON_LOGOUT = Frame(FRAME_MAIN, bg="#ddd")
        BUTTON_LOGOUT = Button(FRAME_BUTTON_LOGOUT, text="LOGOUT", bg="red", font=("arial", 10, "bold"), fg="white", command=self.logout)

        # ---------------------------------------------------#
        # ---------      DISPLAY WIDGET             ---------#
        # ---------------------------------------------------#
        BUTTON_HOME.place(relx=0.01, rely=0.02, relwidth=0.98, relheight=0.96)
        FRAME_BUTTON_HOME.place(relx=0.014, rely=0.44, relwidth=0.122, relheight=0.065)
        LABEL_MENU.place(relx=0, rely=0.03, relwidth=1, relheight=0.94)
        FRAME_MENU.place(relx=0, rely=0.38, relwidth=0.15, relheight=0.05)
        USER_POSITION_LABEL.place(relx=0, rely=0.292, relwidth=0.15, relheight=0.05)
        USER_LOGO_LABEL.place(relx=0, rely=0.09, relwidth=0.15, relheight=0.20)
        FRAME_USER.place(relx=0, rely=0.09, relwidth=0.15, relheight=0.253)
        FRAME_SIDE.place(relx=0, rely=0.09, relwidth=0.15, relheight=1)
        BUTTON_LOGOUT.place(relx=0.01, rely=0.02, relwidth=0.98, relheight=0.96)
        FRAME_BUTTON_LOGOUT.place(relx=0.014, rely=0.815, relwidth=0.122, relheight=0.065)
        BUTTON_USER.place(relx=0.01, rely=0.02, relwidth=0.98, relheight=0.96)
        FRAME_BUTTON_USER.place(relx=0.014, rely=0.74, relwidth=0.122, relheight=0.065)
        BUTTON_INVOICE.place(relx=0.01, rely=0.02, relwidth=0.98, relheight=0.96)
        FRAME_BUTTON_INVOICE.place(relx=0.014, rely=0.665, relwidth=0.122, relheight=0.065)
        BUTTON_INVENTORY.place(relx=0.01, rely=0.02, relwidth=0.98, relheight=0.96)
        FRAME_BUTTON_INVENTORY.place(relx=0.014, rely=0.59, relwidth=0.122, relheight=0.065)
        BUTTON_PRODUCT.place(relx=0.01, rely=0.02, relwidth=0.98, relheight=0.96)
        FRAME_BUTTON_PRODUCT.place(relx=0.014, rely=0.515, relwidth=0.122, relheight=0.065)


    def main_body(self):
        global FRAME_BODY, FRAME_BODY_PADDING
        FRAME_BODY = Frame(FRAME_MAIN, bg="#5cb7d8")
        FRAME_BODY_PADDING = Frame(FRAME_BODY, bg="white")

        # ---- MAIN TOP FRAME ----#
        global FRAME_BODY_TOP
        FRAME_BODY_TOP = Frame(FRAME_BODY_PADDING, bg="#5cb7d8")

        global ICON_PRODUCT, ICON_PRODUCT_LABEL
        ICON_PRODUCT = PhotoImage(file="icon/home.png")
        ICON_PRODUCT_LABEL = Label(FRAME_BODY_TOP, image=ICON_PRODUCT, bg="#5cb7d8")

        global LABEL_SUBTITLE
        LABEL_SUBTITLE = Label(FRAME_BODY_TOP, text="HOME", font=("algerian", 14), bg="#5cb7d8")

        global open_modal, ICON_SEARCH
        ICON_SEARCH = PhotoImage(file="icon/search.gif")

        # ---------------------------------------------------#
        # ---------      DISPLAY WIDGET             ---------#
        # ---------------------------------------------------#
        LABEL_SUBTITLE.grid(row=0, column=0, padx=40, pady=5)
        ICON_PRODUCT_LABEL.place(relx=0.016, rely=0.1)
        FRAME_BODY_TOP.place(relx=0, rely=0, relwidth=1, relheight=0.08)
        FRAME_BODY_PADDING.place(relx=0.004, rely=0.004, relwidth=0.994, relheight=0.993)
        FRAME_BODY.place(relx=0.18, rely=0.12, relwidth=0.79, relheight=0.78)

        # --------- CALLING FUNCTION ---------#
        frame_home.show()
        frame_product.show()
        frame_inventory.show()
        frame_invoice.show()
        frame_user.show()

    def footer(self):
        global FRAME_BOTTOM, LABEL_AUTHOR, LABEL_VERSION
        FRAME_BOTTOM = Frame(FRAME_MAIN, bg="white")
        LABEL_AUTHOR = Label(FRAME_BOTTOM, text="Copyright @ 2019 || SIS. All rights reserved.", bg="white")
        LABEL_VERSION = Label(FRAME_BOTTOM, text="Version: 1.0.", bg="white")
        FRAME_BOTTOM.place(relx=0, rely=0.96, relwidth=1, relheight=0.04)
        LABEL_AUTHOR.pack()
        LABEL_VERSION.place(relx=0.91, y=0)

    def show_frame(self, this_frame):
        # Hide All frame group
        frame_home.hide()
        frame_product.hide()
        frame_inventory.hide()
        frame_invoice.hide()
        frame_user.hide()

        # ------ Button Color reset ----#
        BUTTON_HOME.config(bg="#1d1dab")
        BUTTON_PRODUCT.config(bg="#1d1dab")
        BUTTON_INVOICE.config(bg="#1d1dab")
        BUTTON_USER.config(bg="#1d1dab")
        BUTTON_INVENTORY.config(bg="#1d1dab")

        if this_frame == "home":
            frame_home.display()
            BUTTON_HOME.config(bg="blue")

        elif this_frame == 'product':
            frame_product.display()
            BUTTON_PRODUCT.config(bg="blue")

        elif this_frame == 'inventory':
            frame_inventory.display()
            BUTTON_INVENTORY.config(bg="blue")

        elif this_frame == 'invoice':
            frame_invoice.display()
            BUTTON_INVOICE.config(bg="blue")

        elif this_frame == 'user':
            frame_user.display()
            BUTTON_USER.config(bg="blue")

    def home(self):
        self.show_frame("home")

    def product(self):
        self.show_frame("product")

    def inventory(self):
        self.show_frame("inventory")

    def invoice(self):
        self.show_frame("invoice")

    def user(self):
        self.show_frame("user")

    def close(self):
        if messagebox.askyesno("Confirmation message", "Are you sure you want to close this app?",
                               parent=FRAME_MAIN) == True:
            FRAME_MAIN.destroy()
            ROOT.destroy()

    def logout(self):
        choice = messagebox.askquestion("Confirmation message", "Are you sure you want to logout?", icon='warning',
                                        parent=FRAME_MAIN)
        if choice == "yes":
            FRAME_MAIN.destroy()
            ROOT.update()
            ROOT.deiconify()
            login_cancel()

    def time(self):
        string = strftime('%H:%M:%S %p')
        LABEL_TIME.config(text=string)
        LABEL_TIME.after(1000, self.time)


"""
#====================================================================================================================#
#-----------------------            LOGIN  CLASS                                                -----------------------#
#====================================================================================================================#
"""
class Login:
    def __init__(self):
        pass

    def show(self):
        global ROOT, LOGIN_WIDTH, LOGIN_HEIGHT
        LOGIN_WIDTH = 650
        LOGIN_HEIGHT = 600

        ROOT = Tk()
        config = window_config(ROOT, LOGIN_WIDTH, LOGIN_HEIGHT)

        FRAME_TOP_LOGIN = Frame(ROOT, bg="blue")
        FRAME_TOP_LOGIN.place(relx=0, rely=0, relwidth=1, relheight=0.11)

        FRAME_LOGIN_INNER = Frame(ROOT, bg="#5cb7d8")
        FRAME_LOGIN_INNER.place(relx=0, rely=0, relwidth=1, relheight=0.10)

        global LOGIN_BACKGROUND
        LOGIN_BACKGROUND = PhotoImage(file="icon/system_logo.png")
        LABEL_LOGIN_BACKGROUND = Label(FRAME_LOGIN_INNER, image=LOGIN_BACKGROUND, bg="#5cb7d8")
        LABEL_LOGIN_BACKGROUND.place(x=100, y=0)

        LABEL_LOGIN_TITLE = Label(FRAME_LOGIN_INNER, text="STORE INVENTORY SYSTEM", bg="#5cb7d8", font=("algerian", 20))
        LABEL_LOGIN_TITLE.grid(row=0, column=1, padx=160, ipady=11)

        FRAME_CENTER_LOGIN = Frame(ROOT, bg="#5cb7d8")
        FRAME_CENTER_LOGIN.place(relx=0.15, rely=0.18, relwidth=0.70, relheight=0.70)

        FRAME_CENTER_LOGIN_IMAGE = Frame(ROOT, bg="white")
        FRAME_CENTER_LOGIN_IMAGE.place(relx=0.16, rely=0.19, relwidth=0.68, relheight=0.68)

        global ICON_LOCK_LOGIN
        ICON_LOCK_LOGIN = PhotoImage(file="icon/locker.png")
        LABEL_LOCK_LOGIN = Label(FRAME_CENTER_LOGIN_IMAGE, image=ICON_LOCK_LOGIN, bg="white")
        LABEL_LOCK_LOGIN.grid(row=0, column=0, ipady=11, padx=10)

        LABEL_FORM_LOGIN = Label(FRAME_CENTER_LOGIN_IMAGE, text="Login form: ", font=("algerian", 18), bg="white")
        LABEL_FORM_LOGIN.grid(row=0, column=1, pady=2)

        global LOGO_LOGIN
        LOGO_LOGIN = PhotoImage(file="icon/user.png")
        LABEL_LOGO_LOGIN = Label(FRAME_CENTER_LOGIN_IMAGE, image=LOGO_LOGIN, bg="white")
        LABEL_LOGO_LOGIN.place(x=166, y=60)

        FRAME_FORM_LOGIN = Frame(FRAME_CENTER_LOGIN_IMAGE, bg="white")
        FRAME_FORM_LOGIN.place(x=55, y=180)

        global LABEL_USERNAME_LOGIN
        LABEL_USERNAME_LOGIN = Label(FRAME_FORM_LOGIN, text="Username: ", font=("arial", 13), bg="white")
        LABEL_USERNAME_LOGIN.grid(row=3, column=0, sticky="w")

        global username_login_var, ENTRY_USERNAME_LOGIN
        username_login_var = StringVar()
        ENTRY_USERNAME_LOGIN = Entry(FRAME_FORM_LOGIN, textvariable=username_login_var, font=("arial", 13), bg="#ddd", width=36, bd=2,
                               relief=GROOVE)
        ENTRY_USERNAME_LOGIN.grid(row=4, column=0, ipady=5, padx=5)
        ENTRY_USERNAME_LOGIN.focus_set()

        LABEL_SPACING_LOGIN = Label(FRAME_FORM_LOGIN, font=("arial", 13), bg="white")
        LABEL_SPACING_LOGIN.grid(row=5, column=0, sticky="w")

        LABEL_PASSWORD_LOGIN = Label(FRAME_FORM_LOGIN, text="Password: ", font=("arial", 13), bg="white")
        LABEL_PASSWORD_LOGIN.grid(row=6, column=0, sticky="w")

        global password_login_var, ENTRY_PASSWORD_LOGIN
        password_login_var = StringVar()
        ENTRY_PASSWORD_LOGIN = Entry(FRAME_FORM_LOGIN, textvariable=password_login_var, font=("arial", 13), bg="#ddd", width=36, bd=2,
                               relief=GROOVE, show="*")
        ENTRY_PASSWORD_LOGIN.grid(row=7, column=0, ipady=5)

        FRAME_BUTTON_LOGIN = Frame(FRAME_CENTER_LOGIN_IMAGE, bg="white")
        FRAME_BUTTON_LOGIN.place(x=100, y=335)

        BUTTON_CANCEL_LOGIN = Button(FRAME_BUTTON_LOGIN, text="CANCEL", width=10, bg='red', fg="white", font=("arial", 11), relief=GROOVE, command=login_cancel)
        BUTTON_CANCEL_LOGIN.pack(side=LEFT, pady=15, padx=10, ipady=2)


        BUTTON_SUBMIT_LOGIN = Button(FRAME_BUTTON_LOGIN, text="LOGIN", width=10, bg="blue", fg="white", font=("arial", 11),
                               relief=GROOVE, command=login_submit)
        BUTTON_SUBMIT_LOGIN.pack(side=LEFT, padx=10, ipady=2)

        FRAME_BOTTOM_LOGIN = Frame(ROOT, bg="white")
        FRAME_BOTTOM_LOGIN.place(relx=0, rely=0.96, relwidth=1, relheight=0.04)

        LABEL_COPYRIGHT_LOGIN = Label(FRAME_BOTTOM_LOGIN, text="Copyright @ 2019 || SIS. All rights reserved.", bg="white")
        LABEL_COPYRIGHT_LOGIN.pack()

        LABEL_VERSION_LOGIN = Label(FRAME_BOTTOM_LOGIN, text="Version: 1.0.", bg="white")
        LABEL_VERSION_LOGIN.place(x=550, y=0)

        ROOT.geometry('%dx%d+%d+%d' % (config['WIDTH'], config['HEIGHT'], config['x'], config['y']))
        ROOT.bind("<Return>", login_enter)
        ROOT.mainloop()

    def display(self):
        pass

    def hide(self):
        pass

    def close(self):
        pass
"""
#====================================================================================================================#
#-----------------------            LOGIN FUNCTIONALITY AREA                                  -----------------------#
#====================================================================================================================#
"""
def window_config(frame, WIDTH, HEIGHT):
    frame.title("STORE INVENTORY SYSTEM")
    ws = frame.winfo_screenwidth()
    hs = frame.winfo_screenheight()
    x = (ws / 2) - (WIDTH / 2)
    y = (hs / 2) - (HEIGHT / 2)
    frame.resizable(False, False)
    frame.iconbitmap('icon/system_logo.ico')
    frame.config(bg="#ddd")
    frame.resizable(False, False)

    return {
        "WIDTH": WIDTH,
        "HEIGHT":
            HEIGHT,
        "x": x,
        "y": y
    }

def login_submit():
    login_function()

def login_cancel():
    username_login_var.set("")
    password_login_var.set("")
    ENTRY_USERNAME_LOGIN.focus_set()

def login_enter(event):
    login_function()

def login_validate():
    errors = []
    if username_login_var.get() == "":
        errors.append("Username is required.")

    if password_login_var.get() == "":
        errors.append("Password is required.")

    return errors

def login_function():
    errors = login_validate()

    if len(errors) == 0:
        uname = username_login_var.get()
        upass = password_login_var.get()
        resp = controller.user_login(uname, upass)

        if len(resp) > 0:
            if resp[0]['user_password'] == upass:
                messagebox.showinfo("Information message", "Login successfully.", parent=ROOT)
                ROOT.withdraw()
                frame_main.show()
            else:
                messagebox.showwarning("Failed", "Incorrect username or password", parent=ROOT)
        else:
            messagebox.showwarning("Failed", "Account not Exist", parent=ROOT)
            login_cancel()
    else:
        ENTRY_USERNAME_LOGIN.focus()
        message = ""
        for error in errors:
            message += error + "\n"

        messagebox.showerror("Error message", message, parent=ROOT)





class Home:
    def __init__(self):
        pass

    def show(self):
        # ---  ADD THIS CODE ---- #
        # ----  FORM WIDGET ---#
        global FRAME_HOME_WIDGET, BG_HOME
        BG_HOME = PhotoImage(file="icon/bg.png")
        FRAME_HOME_WIDGET = Label(FRAME_BODY_PADDING, image=BG_HOME, bg="white")
        FRAME_HOME_WIDGET.place(relx=0, rely=0.1, relwidth=1, relheight=0.8)
        # ---  ADD THIS CODE ---- #

    def display(self):
        LABEL_SUBTITLE.config(text="Home")
        global ICON_HOME
        ICON_HOME = PhotoImage(file="icon/home.png")
        ICON_PRODUCT_LABEL.config(image=ICON_HOME)
        # ---  ADD THIS CODE ---- #
        FRAME_HOME_WIDGET.place(relx=0, rely=0.1, relwidth=1, relheight=0.8)
        # ---  ADD THIS CODE ---- #

    def hide(self):
        LABEL_SUBTITLE.config(text="")
        # ---  ADD THIS CODE ---- #
        FRAME_HOME_WIDGET.place_forget()
        # ---  ADD THIS CODE ---- #

    def close(self):
        pass


"""
#====================================================================================================================#
#-----------------------            PRODUCTS CLASS AREA                                       -----------------------#
#====================================================================================================================#
"""
class Product:
    def __init__(self):
        pass

    def show(self):
        # ----  FORM WIDGET ---#
        global FRAME_PRODUCT_WIDGET
        FRAME_PRODUCT_WIDGET = Frame(FRAME_BODY_PADDING, bg="white")

        # stock no
        LABEL_STOCK_NO = Label(FRAME_PRODUCT_WIDGET, text="Stock no:", font=("arial", 11, "bold"), bg="white")

        global stockno_var, ENTRY_STOCK_NO
        stockno_var = StringVar()
        ENTRY_STOCK_NO = Entry(FRAME_PRODUCT_WIDGET, textvariable=stockno_var, font=("arial", 11), width=60, bd=2,
                               relief=GROOVE)
        ENTRY_STOCK_NO.bind("<Return>", product_registration_searchkey)
        ENTRY_STOCK_NO.bind("<KeyRelease>", product_registration_searchkey)
        ENTRY_STOCK_NO.focus()

        BUTTON_SEARCH_STOCKNO = Button(FRAME_PRODUCT_WIDGET, image=ICON_SEARCH, bg="#5cb7d8", font=("arial", 11),
                                       width=30, height=23, bd=2, relief=GROOVE)
        BUTTON_SEARCH_STOCKNO.bind("<Button-1>", button_search_pr)

        # product name
        LABEL_PRODUCT_NAME = Label(FRAME_PRODUCT_WIDGET, text="Product name:", font=("arial", 11, "bold"), bg="white")

        global product_name_var, ENTRY_PRODUCT_NAME
        product_name_var = StringVar()
        ENTRY_PRODUCT_NAME = Entry(FRAME_PRODUCT_WIDGET, textvariable=product_name_var, font=("arial", 11), width=60,
                                   bd=2, relief=GROOVE)

        # products_details
        LABEL_PRODUCT_DETAIL = Label(FRAME_PRODUCT_WIDGET, text="Product details:", font=("arial", 11, "bold"),
                                     bg="white")

        global product_detail_var, ENTRY_PRODUCT_DETAIL
        product_detail_var = StringVar()
        ENTRY_PRODUCT_DETAIL = Entry(FRAME_PRODUCT_WIDGET, textvariable=product_detail_var, font=("arial", 11),
                                     width=60, bd=2, relief=GROOVE)

        # products brand
        LABEL_PRODUCT_BRAND = Label(FRAME_PRODUCT_WIDGET, text="Product brand:", font=("arial", 11, "bold"), bg="white")

        global product_brand_var, ENTRY_PRODUCT_BRAND
        product_brand_var = StringVar()
        ENTRY_PRODUCT_BRAND = Entry(FRAME_PRODUCT_WIDGET, textvariable=product_brand_var, font=("arial", 11), width=60,
                                    bd=2, relief=GROOVE)

        # products price
        LABEL_PRODUCT_PRICE = Label(FRAME_PRODUCT_WIDGET, text="Price:", font=("arial", 11, "bold"), bg="white")

        global product_price_var, ENTRY_PRODUCT_PRICE
        product_price_var = DoubleVar()
        product_price_var.set(0.00)
        ENTRY_PRODUCT_PRICE = Entry(FRAME_PRODUCT_WIDGET, textvariable=product_price_var, font=("arial", 11), width=60,
                                    bd=2, relief=GROOVE)
        ENTRY_PRODUCT_PRICE.bind("<FocusIn>", product_registration_focusprice)
        ENTRY_PRODUCT_PRICE.bind("<KeyPress>", product_registration_numberonly_price)
        ENTRY_PRODUCT_PRICE.bind("<KeyRelease>", product_registration_numberonly_price)

        # products quantity
        LABEL_PRODUCT_QUANTITY = Label(FRAME_PRODUCT_WIDGET, text="Quantity:", font=("arial", 11, "bold"), bg="white")

        global product_quantity_var, ENTRY_PRODUCT_QUANTITY
        product_quantity_var = IntVar()
        ENTRY_PRODUCT_QUANTITY = Spinbox(FRAME_PRODUCT_WIDGET, textvariable=product_quantity_var, from_=0, to=1000,
                                         font=("arial", 11), width=59, bd=2, relief=GROOVE)
        ENTRY_PRODUCT_QUANTITY.bind("<FocusIn>", product_registration_focusqty)
        ENTRY_PRODUCT_QUANTITY.bind("<KeyRelease>", product_registration_numberonly_qty)
        ENTRY_PRODUCT_QUANTITY.bind("<KeyPress>", product_registration_numberonly_qty)

        global message_var
        message_var = StringVar()
        LABEL_MESSAGE = Label(FRAME_PRODUCT_WIDGET, textvariable=message_var, font=("arial", 10, "bold"), fg="blue",
                              bg="white")

        # ---- FRAME CANCEL AND SUBMIT -----#
        global FRAME_PRODUCT_SUBMIT, BUTTON_PRODUCT_CANCEL, BUTTON_PRODUCT_SUBMIT, BUTTON_PRODUCT_UPDATE
        FRAME_PRODUCT_SUBMIT = Frame(FRAME_PRODUCT_WIDGET, bg="white")

        # cancel button
        BUTTON_PRODUCT_UPDATE = Button(FRAME_PRODUCT_SUBMIT, text="UPDATE", width=10, bg='orange', fg="white",
                                       font=("arial", 11), relief=GROOVE, command=update_product)
        #BUTTON_PRODUCT_UPDATE.bind("<Button-1>", product_update)

        # cancel button
        BUTTON_PRODUCT_CANCEL = Button(FRAME_PRODUCT_SUBMIT, text="CANCEL", width=10, bg='red', fg="white",
                                       font=("arial", 11), relief=GROOVE)
        BUTTON_PRODUCT_CANCEL.bind("<Button-1>", product_registration_cancel)

        # submit button
        BUTTON_PRODUCT_SUBMIT = Button(FRAME_PRODUCT_SUBMIT, text="SUBMIT", width=10, bg="blue", fg="white",
                                       font=("arial", 11), relief=GROOVE)
        BUTTON_PRODUCT_SUBMIT.bind("<Button-1>", product_registration_submit)

        # ----------------------------------------------#
        # --------        DISPLAY WIDGET       -------- #
        # ----------------------------------------------#

        # ----Main frame that holds all widget -----
        # FRAME_PRODUCT_WIDGET.place(relx=0.08, rely=0.18, relwidth=1, relheight=1)
        BUTTON_PRODUCT_UPDATE.pack(side=LEFT, padx=10, ipady=2)
        BUTTON_PRODUCT_CANCEL.pack(side=LEFT, padx=10, ipady=2)
        BUTTON_PRODUCT_SUBMIT.pack(side=LEFT, pady=15, padx=10, ipady=2)
        FRAME_PRODUCT_SUBMIT.place(relx=0.2, rely=0.55)
        LABEL_MESSAGE.grid(row=6, column=1, ipady=3, padx=40, sticky="w")
        ENTRY_PRODUCT_QUANTITY.grid(row=5, column=1, ipady=3, padx=40, pady=5)
        LABEL_PRODUCT_QUANTITY.grid(row=5, column=0, sticky="W")
        ENTRY_PRODUCT_PRICE.grid(row=4, column=1, ipady=3, padx=40, pady=5)
        LABEL_PRODUCT_PRICE.grid(row=4, column=0, sticky="W")
        ENTRY_PRODUCT_BRAND.grid(row=3, column=1, ipady=3, padx=40, pady=5)
        LABEL_PRODUCT_BRAND.grid(row=3, column=0, sticky="W")
        ENTRY_PRODUCT_DETAIL.grid(row=2, column=1, ipady=3, padx=40, pady=5)
        LABEL_PRODUCT_DETAIL.grid(row=2, column=0, sticky="W")
        ENTRY_PRODUCT_NAME.grid(row=1, column=1, ipady=3, padx=4, pady=5)
        LABEL_PRODUCT_NAME.grid(row=1, column=0, sticky="W")
        BUTTON_SEARCH_STOCKNO.place(x=610, y=5)
        ENTRY_STOCK_NO.grid(row=0, column=1, ipady=3, padx=4, pady=5)
        LABEL_STOCK_NO.grid(row=0, column=0, sticky="W")

    def display(self):
        FRAME_PRODUCT_WIDGET.place(relx=0.08, rely=0.18, relwidth=1, relheight=1)
        LABEL_SUBTITLE.config(text="ADD STOCK")
        global ICON_ADDSTOCK
        ICON_ADDSTOCK = PhotoImage(file="icon/addstock.png")
        ICON_PRODUCT_LABEL.config(image=ICON_ADDSTOCK)

    def hide(self):
        FRAME_PRODUCT_WIDGET.place_forget()
        LABEL_SUBTITLE.config(text="")

    def modal_update(self):
        # ---- WINDOW TOP CONFIG ----#
        global MODAL_UPDATE
        MODAL_UPDATE = Toplevel(FRAME_MAIN)
        config = window_config(MODAL_UPDATE, 550, 450)

        FRAME_MODAL_UPDATE = Frame(MODAL_UPDATE, bg="white")
        FRAME_MODAL_UPDATE.place(relx=0.06, rely=0.02, relwidth=0.88, relheight=0.94)

        # ----  MODAL USER TOP --- #
        global FRAME_MODAL_CUSTOMER_REGISTER_TOP
        FRAME_MODAL_UPDATE_TOP = Frame(FRAME_MODAL_UPDATE, bg="#7fdb98")
        FRAME_MODAL_UPDATE_TOP.place(relx=0, rely=0, relwidth=1, relheight=0.09)

        global ICON_CUSTOMER_VERIFY
        ICON_CUSTOMER_VERIFY = PhotoImage(file="icon/register.png")
        LABEL_ICON_CUSTOMER_VERIFY = Label(FRAME_MODAL_UPDATE_TOP, image=ICON_CUSTOMER_VERIFY, bg="#7fdb98")
        LABEL_ICON_CUSTOMER_VERIFY.place(relx=0.21, rely=0.01, relwidth=0.08, relheight=1)

        LABEL_TOP_TITLE = Label(FRAME_MODAL_UPDATE_TOP, text="UPDATE PRODUCT", font=("algerian", 18),
                                bg="#7fdb98")
        LABEL_TOP_TITLE.place(relx=0.28, rely=0.5, relwidth=0.48, relheight=1, anchor="w")

        # ---- MODAL USER BODY ---- #
        FRAME_MODAL_UPDATE_BODY = Frame(FRAME_MODAL_UPDATE, bg="white")
        FRAME_MODAL_UPDATE_BODY.place(relx=0, rely=0.16, relwidth=0.9
                                      , relheight=0.8)

        FRAME_UPDATE_FORM = Frame(FRAME_MODAL_UPDATE_BODY, bg="white")
        FRAME_UPDATE_FORM.place(relx=0.1, rely=0.05, relwidth=1)

        # ---- PRODUCT STOCKNO UPDATE ----- #
        global LABEL_PRODUCT_STOCKNO_UPDATE
        LABEL_PRODUCT_STOCKNO_UPDATE = Label(FRAME_UPDATE_FORM, text="Stock no: ", font=("arial", 10, "bold"),
                                          bg="white")
        LABEL_PRODUCT_STOCKNO_UPDATE.grid(row=0, column=0, sticky="W")

        global ENTRY_PRODUCT_STOCK_UPDATE, product_stockno_update_var
        product_stockno_update_var = StringVar()
        ENTRY_PRODUCT_STOCK_UPDATE = Entry(FRAME_UPDATE_FORM, textvariable=product_stockno_update_var,
                                          font=("arial", 10), width=38,
                                          bd=2, state="disabled", relief=GROOVE)
        ENTRY_PRODUCT_STOCK_UPDATE.grid(row=0, column=1, ipady=3, pady=5)

        # ---- PRODUCT NAME UPDATE ----- #
        global LABEL_PRODUCT_NAME_UPDATE
        LABEL_PRODUCT_NAME_UPDATE = Label(FRAME_UPDATE_FORM, text="Product name: ", font=("arial", 10, "bold"),
                                      bg="white")
        LABEL_PRODUCT_NAME_UPDATE.grid(row=1, column=0, sticky="W")

        global ENTRY_PRODUCT_NAME_UPDATE, product_name_update_var
        product_name_update_var = StringVar()
        ENTRY_PRODUCT_NAME_UPDATE = Entry(FRAME_UPDATE_FORM, textvariable=product_name_update_var,
                                      font=("arial", 10), width=38,
                                      bd=2, relief=GROOVE)
        ENTRY_PRODUCT_NAME_UPDATE.grid(row=1, column=1, ipady=3, pady=5)

        # ---- PRODUCT DETAILS UPDATE ----- #
        LABEL_PRODUCT_DETAILS_UPDATE = Label(FRAME_UPDATE_FORM, text="Product details:", font=("arial", 10, "bold"),
                                        bg="white")
        LABEL_PRODUCT_DETAILS_UPDATE.grid(row=2, column=0, sticky="W")

        global product_details_update_var, ENTRY_PRODUCT_DETAILS_UPDATE
        product_details_update_var = StringVar()
        ENTRY_PRODUCT_DETAILS_UPDATE = Entry(FRAME_UPDATE_FORM, textvariable=product_details_update_var,
                                        font=("arial", 10), width=38,
                                        bd=2, relief=GROOVE)
        ENTRY_PRODUCT_DETAILS_UPDATE.grid(row=2, column=1, ipady=3, pady=5)

        # ---- PRODUCT BRAND UPDATE ----- #
        LABEL_PRODUCT_BRAND_UPDATE = Label(FRAME_UPDATE_FORM, text="Product brand:", font=("arial", 10, "bold"),
                                             bg="white")
        LABEL_PRODUCT_BRAND_UPDATE.grid(row=3, column=0, sticky="W")

        global product_brand_update_var, ENTRY_PRODUCT_BRAND_UPDATE
        product_brand_update_var = StringVar()
        ENTRY_PRODUCT_BRAND_UPDATE = Entry(FRAME_UPDATE_FORM, textvariable=product_brand_update_var,
                                       font=("arial", 10), width=38,
                                       bd=2, relief=GROOVE)
        ENTRY_PRODUCT_BRAND_UPDATE.grid(row=3, column=1, ipady=3, pady=5)

        # ---- FRAME CANCEL AND SUBMIT -----#
        global FRAME_CUSTOMER_BUTTONS_UPDATE
        FRAME_CUSTOMER_BUTTONS_UPDATE = Frame(FRAME_MODAL_UPDATE, bg="white")
        FRAME_CUSTOMER_BUTTONS_UPDATE.place(relx=0.35, rely=0.6)


        global BUTTON_PRODUCT_SUBMIT_UPDATE, BUTTON_PRODUCT_CANCEL_UPDATE
        BUTTON_PRODUCT_CANCEL_UPDATE = Button(FRAME_CUSTOMER_BUTTONS_UPDATE, text="CANCEL", width=10, bg='red', fg="white",
                                    font=("arial", 11), relief=GROOVE)
        BUTTON_PRODUCT_CANCEL_UPDATE.pack(side=LEFT, pady=15, padx=10, ipady=2)
        BUTTON_PRODUCT_CANCEL_UPDATE.bind("<Button-1>", product_update_cancel)

        # submit button
        BUTTON_PRODUCT_SUBMIT_UPDATE = Button(FRAME_CUSTOMER_BUTTONS_UPDATE, text="SUBMIT", width=10, bg="blue", fg="white",
                                  font=("arial", 11), relief=GROOVE)
        BUTTON_PRODUCT_SUBMIT_UPDATE.pack(side=LEFT, padx=10, ipady=2)
        BUTTON_PRODUCT_SUBMIT_UPDATE.bind("<Button-1>", product_update_submit)

        MODAL_UPDATE.grab_set()
        MODAL_UPDATE.geometry('%dx%d+%d+%d' % (config['WIDTH'], config['HEIGHT'], config['x'], config['y']))


    def modal(self):
        # ---- WINDOW TOP CONFIG ----#
        global MODAL_SEARCH
        MODAL_SEARCH = Toplevel(FRAME_MAIN)
        config = window_config(MODAL_SEARCH, 600, 450)

        FRAME_MODAL_SEARCH = Frame(MODAL_SEARCH)
        FRAME_MODAL_SEARCH.place(relx=0.06, rely=0.02, relwidth=0.88, relheight=0.94)

        ICON_SEARCH_PRODUCT = PhotoImage(file="icon/search_product.png").subsample(2, 2)
        LABEL_SEARCH_PRODUCT = Label(MODAL_SEARCH, image=ICON_SEARCH_PRODUCT)
        LABEL_SEARCH_PRODUCT.place(relx=0.36, rely=0.07, relheight=0.28, relwidth=0.27)

        FRAME_MODAL_FORM = Frame(FRAME_MODAL_SEARCH)
        FRAME_MODAL_FORM.place(relx=0, rely=0.35, relheight=1, relwidth=1)

        global modal_stockno
        modal_stockno = StringVar()
        ENTRY_MODAL_STOCKNO = Entry(FRAME_MODAL_FORM, textvariable=modal_stockno, font=("arial", 11), width=43, bd=2,
                                    relief=GROOVE, bg="#ddd")
        ENTRY_MODAL_STOCKNO.grid(row=0, column=1, ipady=6, padx=65, pady=5)
        ENTRY_MODAL_STOCKNO.focus()
        ENTRY_MODAL_STOCKNO.bind("<KeyRelease>", modal_search_stockno)

        BUTTON_MODAL_SEARCH_STOCKNO = Button(FRAME_MODAL_FORM, image=ICON_SEARCH, bg="#5cb7d8", font=("arial", 11),
                                             width=36, height=29,
                                             bd=2,
                                             relief=GROOVE)
        BUTTON_MODAL_SEARCH_STOCKNO.place(x=413, y=5)
        BUTTON_MODAL_SEARCH_STOCKNO.bind("<Button-1>", modal_search_stockno)

        # ---- TABLE FRAME ----#
        FRAME_TABLE_SEARCH = Frame(FRAME_MODAL_FORM)
        FRAME_TABLE_SEARCH.place(relx=0.02, rely=0.14, relwidth=0.96, relheight=0.42)

        global TREE_MODAL_SEARCH
        TREE_MODAL_SEARCH = ttk.Treeview(FRAME_TABLE_SEARCH, selectmode='browse')
        TREE_MODAL_SEARCH.pack(side='left')

        MODAL_SCROLLBAR = ttk.Scrollbar(FRAME_TABLE_SEARCH, orient="vertical", command=TREE_MODAL_SEARCH.yview)
        MODAL_SCROLLBAR.pack(side='right', fill='y')

        TREE_MODAL_SEARCH.configure(yscrollcommand=MODAL_SCROLLBAR.set)
        TREE_MODAL_SEARCH["columns"] = ("one", "two", "three", "four")
        TREE_MODAL_SEARCH.column("#0", width=80)
        TREE_MODAL_SEARCH.column("one", width=90)
        TREE_MODAL_SEARCH.column("two", width=200)
        TREE_MODAL_SEARCH.column("three", width=50)
        TREE_MODAL_SEARCH.column("four", width=50)

        TREE_MODAL_SEARCH.heading("#0", text="Stock no", anchor="w")
        TREE_MODAL_SEARCH.heading("one", text="Product name", anchor="w")
        TREE_MODAL_SEARCH.heading("two", text="Brand name", anchor="w")
        TREE_MODAL_SEARCH.heading("three", text="Price", anchor="w")
        TREE_MODAL_SEARCH.heading("four", text="Quantity", anchor="w")

        TREE_MODAL_SEARCH.place(relx=0, rely=0, relwidth=0.963, relheight=1)
        TREE_MODAL_SEARCH.bind("<Double-1>", selected_item_modal)
        TREE_MODAL_SEARCH.bind("<Return>", selected_item_modal)

        # ----------------------------------------------#
        # --------        DISPLAY WIDGET       -------- #
        # ----------------------------------------------#

        MODAL_SEARCH.grab_set()
        MODAL_SEARCH.bind_all('<Escape>', closing_modal_pr_key)
        MODAL_SEARCH.bind_all('<Tab>', tab_modal_pr)
        MODAL_SEARCH.geometry('%dx%d+%d+%d' % (config['WIDTH'], config['HEIGHT'], config['x'], config['y']))
        MODAL_SEARCH.protocol("WM_DELETE_WINDOW", closing_modal_pr_window)
        MODAL_SEARCH.mainloop()


"""
#====================================================================================================================#
#-----------------------            PRODUCTS FUNCTIONALITY AREA                               -----------------------#
#====================================================================================================================#
"""

def update_product():
    if stockno_var.get() == "":
        messagebox.showwarning("Warning", "No Data to Update", parent=FRAME_MAIN)
        return False

    frame_product.modal_update()

    ps = stockno_var.get()
    pn = product_name_var.get()
    pd = product_detail_var.get()
    pb = product_brand_var.get()


    product_stockno_update_var.set(ps)
    product_name_update_var.set(pn)
    product_details_update_var.set(pd)
    product_brand_update_var.set(pb)


def product_update_cancel(event):
    MODAL_UPDATE.destroy()

def product_update_submit(event):
    ps = product_stockno_update_var.get()
    pn = product_name_update_var.get()
    pd = product_details_update_var.get()
    pb = product_brand_update_var.get()
    res = controller.update_product(ps, pn, pd, pb)
    print("update result: ", res)
    if res == 1:
        messagebox.showinfo("Success", "Data Updated", parent=MODAL_UPDATE)
        product_registration_searchkey(event)
    else:
        messagebox.showerror("Failed", "Can't update data", parent=MODAL_UPDATE)


def product_registration_searchkey(event):
    stock = stockno_var.get()
    if stock != "":
        result = controller.check_stock_no(stock)
        if len(result) > 0:
            ENTRY_PRODUCT_NAME.configure(state="readonly")
            ENTRY_PRODUCT_DETAIL.configure(state="readonly")
            ENTRY_PRODUCT_BRAND.configure(state="readonly")
            ENTRY_PRODUCT_PRICE.configure(state="readonly")
            product_name_var.set(result[0]['product_name'])
            product_detail_var.set(result[0]['product_details'])
            product_brand_var.set(result[0]['product_brand'])
            product_price_var.set(result[0]['stock_price'])
            message_var.set("Type quantity to add in stock.")

        else:
            ENTRY_PRODUCT_NAME.configure(state="normal")
            ENTRY_PRODUCT_DETAIL.configure(state="normal")
            ENTRY_PRODUCT_BRAND.configure(state="normal")
            ENTRY_PRODUCT_PRICE.configure(state="normal")
            product_name_var.set("")
            product_detail_var.set("")
            product_brand_var.set("")
            product_price_var.set(0.00)
            product_quantity_var.set(0)
            message_var.set("")

    # messagebox.showinfo(title="Waring", message="focus out {}".format(e))



def product_registration_focusqty(event):
    try:
        if product_quantity_var.get() == 0:
            product_quantity_var.set("")
    except:
        pass

def product_registration_focusprice(event):
    try:
        if product_price_var.get() == 0.0:
            product_price_var.set("")
    except:
        pass

def product_registration_numberonly_qty(event):
    global product_quantity_var
    numbersonly(product_quantity_var)

def product_registration_numberonly_price(event):
    global product_price_var
    numbersonly(product_price_var)

def numbersonly(entry):
    try:
        value = int(entry.get())
    except:
        entry.set("")


def product_registration_cancel(event):
    stockno_var.set("")
    product_name_var.set("")
    product_detail_var.set("")
    product_brand_var.set("")
    product_price_var.set(0.00)
    product_quantity_var.set(0)
    message_var.set("")



def validation(this_frame):
    err = []
    err_count = 0
    for frame in this_frame.winfo_children():
        if frame.winfo_class() in ["Entry", "Spinbox"]:
            if frame.get() in ["", "0", "0.0"]:
                err_count += 1
                frame.config(bg="#fda6a4")
            else:
                frame.config(bg="white")
    if err_count > 0:
        err.append("There {} Field/s Required to be Filled".format(err_count))
    return err


def product_registration_submit(event):
    errors = validation(FRAME_PRODUCT_WIDGET)


    if len(errors) == 0:
        files = open("_adminsession.txt", 'r')
        data = files.readlines()
        id = int(data[2].split(": ")[1])
        files.close()

        product_id = stockno_var.get()
        product = (product_id, product_name_var.get(), product_detail_var.get(), product_brand_var.get())
        storage = (product_quantity_var.get(), product_price_var.get(), date.today(), product_id, id)
        resp, result = controller.product_add(product_id, product, storage)

        if result != "":
            messagebox.showinfo("Information message", "Success to save data")
            ENTRY_PRODUCT_NAME.configure(state="normal")
            ENTRY_PRODUCT_DETAIL.configure(state="normal")
            ENTRY_PRODUCT_BRAND.configure(state="normal")
            ENTRY_PRODUCT_PRICE.configure(state="normal")
            stockno_var.set("")
            product_name_var.set("")
            product_detail_var.set("")
            product_brand_var.set("")
            product_price_var.set(0.00)
            product_quantity_var.set(0)
            message_var.set("")
        else:
            messagebox.showerror("Error message", "Failed to add data.")
    else:
        messagebox.showerror("Error message", errors[0], parent=FRAME_MAIN)

def product_registration_validate():
    errors = []
    if stockno_var.get() == "":
        errors.append("Stock no is required.")

    if product_name_var.get() == "":
        errors.append("Product name is required.")

    if product_detail_var.get() == "":
        errors.append("Product details is required.")

    if product_brand_var.get() == "":
        errors.append("Product brand is required.")

    try:
        if product_price_var.get() < 0:
            errors.append("You have entered invalid price.")
    except:
        messagebox.showerror("Error message", "You have entered invalid quantity.")

    try:
        if product_quantity_var.get() < 0:
            errors.append("You have entered invalid quantity.")
    except:
        messagebox.showerror("Error message", "You have entered invalid quantity.")

    return errors

def button_search_pr(event):
    frame_product.modal()

def modal_search_stockno(event):

    global product_list
    product_list = {}
    search_val = modal_stockno.get()
    result = controller.search_stock_no(search_val)
    product_list.clear()

    for row in TREE_MODAL_SEARCH.get_children():
        TREE_MODAL_SEARCH.delete(row)

    if len(result) > 0:
        for row in result:
            product_list[row['product_id']] = [{
                'name': row['product_name'],
                'brand': row['product_brand'],
                'details': row['product_details'],
                'price': row['stock_price'],
                'quant': row['stock_quan']
            }]
            TREE_MODAL_SEARCH.insert('', 'end', text=row['product_id'], values=[row['product_name'], row['product_brand'],
                                                                            row['stock_price'], row['stock_quan']])

    else:
        TREE_MODAL_SEARCH.insert('', 'end', text="", values=["No Data Found", "", "", ""])


def selected_item_modal(event):

    item = TREE_MODAL_SEARCH.identify("item", event.x, event.y)
    prod_id = TREE_MODAL_SEARCH.item(item)['text']

    if prod_id == "":
        return False

    MODAL_SEARCH.destroy()
    prod = product_list[prod_id]

    prod_name = prod[0]['name']
    prod_brand = prod[0]['brand']
    prod_price = prod[0]['price']
    prod_detail = prod[0]['details']

    ENTRY_PRODUCT_NAME.configure(state="readonly")
    ENTRY_PRODUCT_DETAIL.configure(state="readonly")
    ENTRY_PRODUCT_BRAND.configure(state="readonly")
    ENTRY_PRODUCT_PRICE.configure(state="readonly")

    stockno_var.set(prod_id)
    product_name_var.set(prod_name)
    product_detail_var.set(prod_detail)
    product_brand_var.set(prod_brand)
    product_price_var.set(prod_price)
    message_var.set("Type quantity to add in stock.")

def tab_modal_pr(event):
    pass

def closing_modal_pr_key(event):
    open_modal = False
    MODAL_SEARCH.destroy()

def closing_modal_pr_window():
    MODAL_SEARCH.destroy()

"""
#====================================================================================================================#
#-----------------------            INVENTORY CLASS AREA                                      -----------------------#
#====================================================================================================================#
"""
class Inventory:
    def __init__(self):
        pass

    def show(self):
        global FRAME_INVENTORY_WIDGET
        FRAME_INVENTORY_WIDGET = Frame(FRAME_BODY_PADDING, bg="white")

        # search
        global LABEL_SEARCH_INVENTORY
        LABEL_SEARCH_INVENTORY = Label(FRAME_INVENTORY_WIDGET, text="Search:", font=("arial", 11, "bold"), bg="white")

        global SEARCH_INVENTORY_VAR, ENTRY_SEARCH_INVENTORY
        SEARCH_INVENTORY_VAR = StringVar()
        ENTRY_SEARCH_INVENTORY = Entry(FRAME_INVENTORY_WIDGET, textvariable=SEARCH_INVENTORY_VAR, bg="#ddd",
                                       font=("arial", 11), width=30, bd=2, relief=GROOVE)
        ENTRY_SEARCH_INVENTORY.focus()
        ENTRY_SEARCH_INVENTORY.bind("<KeyRelease>", inventory_search)

        global BUTTON_SEARCH_INVENTORY
        BUTTON_SEARCH_INVENTORY = Button(FRAME_INVENTORY_WIDGET, image=ICON_SEARCH, bg="#5cb7d8", font=("arial", 11),
                                         width=5, bd=2,
                                         relief=GROOVE)
        BUTTON_SEARCH_INVENTORY.bind("<Button-1>", inventory_search)

        # ---- REFRESH INVENTORY RECORDS ----#
        global ICON_REFRESH_INVENTORY
        ICON_REFRESH_INVENTORY = PhotoImage(file="icon/refresh.png")
        BUTTON_REFRESH_INVENTORY = Button(FRAME_INVENTORY_WIDGET, image=ICON_REFRESH_INVENTORY, bg="#5cb7d8", font=("arial", 11),
                                      width=5, bd=2, relief=GROOVE)

        BUTTON_REFRESH_INVENTORY.place(relx=0.48, rely=0.009, relwidth=0.05, relheight=0.061)
        BUTTON_REFRESH_INVENTORY.bind("<Button-1>", inventory_refresh)

        # ---- EXPORT EXCEL INVENTORY RECORDS ----#
        global ICON_EXCEL_INVENTORY, BUTTON_EXCEL_INVENTORY, is_inventory_selected
        is_inventory_selected = False
        ICON_EXCEL_INVENTORY = PhotoImage(file="icon/excel.png")
        BUTTON_EXCEL_INVENTORY = Button(FRAME_INVENTORY_WIDGET, image=ICON_EXCEL_INVENTORY, bg="#5cb7d8", font=("arial", 11),
                                      bd=2, relief=GROOVE)
        BUTTON_EXCEL_INVENTORY.place(relx=0.54, rely=0.009, relwidth=0.05, relheight=0.061)
        BUTTON_EXCEL_INVENTORY.bind("<Button-1>", inventory_excel)

        # ---- TABLE FRAME ----#
        global FRAME_TABLE_INVENTORY, TREEVIEW_INVENTORY, TREEVIEW_SCROLL_INVENTORY
        FRAME_TABLE_INVENTORY = Frame(FRAME_INVENTORY_WIDGET)

        TREEVIEW_INVENTORY = ttk.Treeview(FRAME_TABLE_INVENTORY, selectmode='browse')

        TREEVIEW_SCROLL_INVENTORY = ttk.Scrollbar(FRAME_TABLE_INVENTORY, orient="vertical",
                                                  command=TREEVIEW_INVENTORY.yview)

        TREEVIEW_INVENTORY.configure(yscrollcommand=TREEVIEW_SCROLL_INVENTORY.set)
        TREEVIEW_INVENTORY["columns"] = ("one", "two", "three", "four", "five", "six")
        TREEVIEW_INVENTORY.column("#0", width=80)
        TREEVIEW_INVENTORY.column("one", width=200)
        TREEVIEW_INVENTORY.column("two", width=50)
        TREEVIEW_INVENTORY.column("three", width=50)
        TREEVIEW_INVENTORY.column("four", width=50)
        TREEVIEW_INVENTORY.column("five", width=30)
        TREEVIEW_INVENTORY.column("six", width=30)

        TREEVIEW_INVENTORY.heading("#0", text="Stock no", anchor="w")
        TREEVIEW_INVENTORY.heading("one", text="Product name", anchor="w")
        TREEVIEW_INVENTORY.heading("two", text="Brand name", anchor="w")
        TREEVIEW_INVENTORY.heading("three", text="Price", anchor="w")
        TREEVIEW_INVENTORY.heading("four", text="Remaining qty", anchor="w")
        TREEVIEW_INVENTORY.heading("five", text="Total qty", anchor="w")
        TREEVIEW_INVENTORY.heading("six", text="Sales qty", anchor="w")
        inventory_load_records("")
        #TREEVIEW_INVENTORY.insert('', 'end', text="---", values=["No Data Found", "---", "---", "---", "---", "---"])

        # ----------------------------------------------#
        # --------        DISPLAY WIDGET       -------- #
        # ----------------------------------------------#

        # -----Main Frame holds all Widget------#
        # FRAME_INVENTORY_WIDGET.place(relx=0.02, rely=0.14, relwidth=1, relheight=1)
        LABEL_SEARCH_INVENTORY.grid(row=0,
                                    column=0, sticky="W")
        ENTRY_SEARCH_INVENTORY.grid(row=0, column=1, ipady=5, padx=4, pady=5)
        BUTTON_SEARCH_INVENTORY.place(relx=0.42, rely=0.009, relwidth=0.05, relheight=0.061)
        FRAME_TABLE_INVENTORY.place(relx=0, rely=0.1, relwidth=0.96, relheight=0.73)
        TREEVIEW_SCROLL_INVENTORY.pack(side='right', fill='y')
        TREEVIEW_INVENTORY.place(relx=0, rely=0, relwidth=0.975, relheight=1)


    def display(self):
        FRAME_INVENTORY_WIDGET.place(relx=0.02, rely=0.14, relwidth=1, relheight=1)
        LABEL_SUBTITLE.config(text="Inventory")
        global ICON_INVENTORY
        ICON_INVENTORY = PhotoImage(file="icon/inventory.png")
        ICON_PRODUCT_LABEL.config(image=ICON_INVENTORY)

    def hide(self):
        FRAME_INVENTORY_WIDGET.place_forget()
        LABEL_SUBTITLE.config(text="")


"""
#====================================================================================================================#
#-----------------------            INVENTORY FUNCTIONALITY AREA                              -----------------------#
#====================================================================================================================#
"""

def inventory_load_records(search_val):
    global excel_inventory_record
    for row in TREEVIEW_INVENTORY.get_children():
        TREEVIEW_INVENTORY.delete(row)

    rows, sales = controller.search_inventory(search_val)
    array = {}

    for row in rows:
        array[row['product_id']] = [{
            'name': row['product_name'],
            'details': row['product_details'],
            'brand': row['product_brand'],
            'price': row['stock_price'],
            'qnty': row['total_qty'],
            'sale': 0,
            'qnty_left': row['total_qty']
        }]

    for sale in sales:
        left = int(array[sale['product_id']][0]['qnty']) - int(sale['sale_quan'])
        array[sale['product_id']][0].update({'sale': sale['sale_quan'], 'qnty_left': left})

    global excel_inventory_record
    excel_inventory_record = array

    if len(array) > 0:
        for row in array:
            TREEVIEW_INVENTORY.insert('', 'end', text=row,
                                        values=[array[row][0]['name'], array[row][0]['brand'], array[row][0]['price'],
                                                array[row][0]['qnty_left'], array[row][0]['qnty'], array[row][0]['sale']])
    else:
        TREEVIEW_INVENTORY.insert('', 'end', text="---", values=["No Data Found", "---", "---", "---", "---", "---"])

def inventory_search(event):
    search_val = SEARCH_INVENTORY_VAR.get()
    inventory_load_records(search_val)

def inventory_refresh(event):
    SEARCH_INVENTORY_VAR.set("")
    inventory_load_records("")

def inventory_excel(event):
    global excel_inventory_record
    try:
        with OpenKey(HKEY_CURRENT_USER, 'SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders') as key:
            downloads_dir = QueryValueEx(key, '{374DE290-123F-4565-9164-39C4925E467B}')[0]

        wb = openpyxl.Workbook()
        sheet = wb.active

        # ---  ROW 1 --- #
        sheet.merge_cells('A1:C1')
        sheet.cell(row=1, column=1).value = "Inventory report"
        sheet.cell(row=1, column=1).font = Font(size=20)

        # ---  ROW 3 --- #
        sheet.cell(row=3, column=1).value = "Date: "
        sheet.cell(row=3, column=1).font = Font(size=12, bold=True)

        sheet.cell(row=3, column=2).value = "10/10/2019"
        sheet.cell(row=3, column=2).font = Font(size=12)


        sheet.merge_cells('A{}:B{}'.format(5, 5))
        sheet.cell(row=5, column=1).value = "Stock no"
        sheet.cell(row=5, column=1).font = Font(size=12, bold=True)

        sheet.merge_cells('C{}:D{}'.format(5, 5))
        sheet.cell(row=5, column=3).value = "Product name"
        sheet.cell(row=5, column=3).font = Font(size=12, bold=True)
        #

        sheet.merge_cells('E{}:G{}'.format(5, 5))
        sheet.cell(row=5, column=5).value = "Product details"
        sheet.cell(row=5, column=5).font = Font(size=12, bold=True)

        sheet.cell(row=5, column=8).value = "Brand"
        sheet.cell(row=5, column=8).font = Font(size=12, bold=True)

        sheet.cell(row=5, column=9).value = "Price"
        sheet.cell(row=5, column=9).font = Font(size=12, bold=True)

        sheet.cell(row=5, column=10).value = "Remaining qty"
        sheet.cell(row=5, column=10).font = Font(size=12, bold=True)

        sheet.cell(row=5, column=11).value = "Total qty"
        sheet.cell(row=5, column=11).font = Font(size=12, bold=True)

        sheet.cell(row=5, column=12).value = "Sales qty"
        sheet.cell(row=5, column=12).font = Font(size=12, bold=True)

        row_current = 6

        for row in excel_inventory_record:
            sheet.merge_cells('A{}:B{}'.format(row_current, row_current))
            sheet.cell(row=row_current, column=1).value = row
            sheet.cell(row=row_current, column=1).font = Font(size=12)

            sheet.merge_cells('C{}:D{}'.format(row_current, row_current))
            sheet.cell(row=row_current, column=3).value = excel_inventory_record[row][0]['name']
            sheet.cell(row=row_current, column=3).font = Font(size=12)

            sheet.merge_cells('E{}:G{}'.format(row_current, row_current))
            sheet.cell(row=row_current, column=5).value = excel_inventory_record[row][0]['details']
            sheet.cell(row=row_current, column=5).font = Font(size=12)

            sheet.cell(row=row_current, column=8).value = excel_inventory_record[row][0]['brand']
            sheet.cell(row=row_current, column=8).font = Font(size=12)

            sheet.cell(row=row_current, column=9).value = "{:.2f}".format(excel_inventory_record[row][0]['price'])
            sheet.cell(row=row_current, column=9).font = Font(size=12)

            sheet.cell(row=row_current, column=10).value = excel_inventory_record[row][0]['qnty_left']
            sheet.cell(row=row_current, column=10).font = Font(size=12)
            sheet.cell(row=row_current, column=10).alignment = Alignment(horizontal='center')

            sheet.cell(row=row_current, column=11).value = excel_inventory_record[row][0]['qnty']
            sheet.cell(row=row_current, column=11).alignment = Alignment(horizontal='center')

            sheet.cell(row=row_current, column=12).value =excel_inventory_record[row][0]['sale']
            sheet.cell(row=row_current, column=12).font = Font(size=12)
            sheet.cell(row=row_current, column=12).alignment = Alignment(horizontal='center')
            row_current += 1


        wb.save('{}/{}'.format(downloads_dir, "inventory_report.xlsx"))
        messagebox.showinfo("Information message", "Inventory report was succeffully saved in download folder.")
    except:
        messagebox.showerror("Error message", "Inventory report not saved.")

"""
#====================================================================================================================#
#-----------------------            RECORDS CLASS AREA                                        -----------------------#
#====================================================================================================================#
"""
class Invoice:
    def __init__(self):
        pass

    def show(self):
        # ---- FORM FRAME ----#
        global FRAME_PRODUCT_INVOICE
        FRAME_PRODUCT_INVOICE = Frame(FRAME_BODY_PADDING, bg="white")

        # ---- TAB FRAME ----#
        global NOTEBOOK_INVOICE_STYLE, NOTEBOOK_INVOICE
        NOTEBOOK_INVOICE_STYLE = ttk.Style()
        NOTEBOOK_INVOICE_STYLE.configure('TNotebook.Tab', font=('arial', '10', 'bold'))
        NOTEBOOK_INVOICE_STYLE.configure("TNotebook.Tab", padding=[5, 5])
        NOTEBOOK_INVOICE_STYLE.map("TNotebook.Tab", background=[("selected", "blue")])
        NOTEBOOK_INVOICE = ttk.Notebook(FRAME_PRODUCT_INVOICE)

        global TAB_INVOICE_INVOICE, TAB_INVOICE_SALE, TAB_INVOICE_CUSTOMER
        TAB_INVOICE_INVOICE = Frame(NOTEBOOK_INVOICE, bg="white")
        NOTEBOOK_INVOICE.add(TAB_INVOICE_INVOICE, text="Invoice")

        TAB_INVOICE_SALE = Frame(NOTEBOOK_INVOICE, bg="white")
        NOTEBOOK_INVOICE.add(TAB_INVOICE_SALE, text="Sales")

        TAB_INVOICE_CUSTOMER = Frame(NOTEBOOK_INVOICE, bg="white")
        NOTEBOOK_INVOICE.add(TAB_INVOICE_CUSTOMER, text="Customer info")
        NOTEBOOK_INVOICE.bind("<KeyRelease>", search_sales)

        # -----------------------------------------------#
        # --------  INVOICE  FRAME FORM WIDGET  -------- #
        # ----------------------------------------------#
        global FORM_INVOICE
        FORM_INVOICE = Frame(TAB_INVOICE_INVOICE, bg="#ddd")

        global FORM_INVOICE_WIDGET, LABEL_INVOICE_TITLE
        FORM_INVOICE_WIDGET = Frame(FORM_INVOICE, bg="white")
        FRAME_INVOICE_TITLE = Frame(FORM_INVOICE_WIDGET, bg="#ddd")
        LABEL_INVOICE_TITLE = Label(FORM_INVOICE_WIDGET, text="Search invoice:", font=('arial', '10', 'bold'),
                                    bg="#ddd")

        global FRAME_TAB_INVOICE, LABEL_MARGIN_INVOICE
        FRAME_TAB_INVOICE = Frame(FORM_INVOICE_WIDGET, bg="white")
        LABEL_MARGIN_INVOICE = Label(FRAME_TAB_INVOICE, text="", bg="white")

        # ---- SEARCH INVOICE RECORDS ----#
        global SEARCH_INVOICE_VAR, ENTRY_SEARCH_INVOICE
        SEARCH_INVOICE_VAR = StringVar()
        ENTRY_SEARCH_INVOICE = Entry(FRAME_TAB_INVOICE, textvariable=SEARCH_INVOICE_VAR, bg="#ddd", font=('arial', '11', 'normal'),
                                     bd=2, width=37, relief=GROOVE)
        ENTRY_SEARCH_INVOICE.bind("<KeyRelease>", invoice_search)

        global ICON_SEARCH_INVOICE, BUTTON_SEARCH_INVOICE
        ICON_SEARCH_INVOICE = PhotoImage(file="icon/search.gif")
        BUTTON_SEARCH_INVOICE = Button(FRAME_TAB_INVOICE, image=ICON_SEARCH_INVOICE, bg="#5cb7d8", font=("arial", 11),
                                       bd=2, relief=GROOVE)
        BUTTON_SEARCH_INVOICE.bind("<Button-1>", invoice_search)

        # ---- REFRESH INVOICE RECORDS ----#
        global ICON_REFRESH_INVOICE
        ICON_REFRESH_INVOICE = PhotoImage(file="icon/refresh.png")
        BUTTON_REFRESH_INVOICE = Button(FORM_INVOICE_WIDGET, image=ICON_REFRESH_INVOICE, bg="white", font=("arial", 11),
                                      bd=2, relief=GROOVE)
        BUTTON_REFRESH_INVOICE.bind("<Button-1>", refresh_invoice)

        # ---- EXPORT EXCEL INVOICE RECORDS ----#
        global ICON_EXCEL_INVOICE, BUTTON_EXCEL_INVOICE, is_invoice_selected
        is_invoice_selected = False
        ICON_EXCEL_INVOICE = PhotoImage(file="icon/excel.png")
        BUTTON_EXCEL_INVOICE = Button(FORM_INVOICE_WIDGET, image=ICON_EXCEL_INVOICE, bg="white", font=("arial", 11),
                                       bd=2, relief=GROOVE)
        BUTTON_EXCEL_INVOICE.bind("<Button-1>", invoice_excel)

        # ---- TABLE FRAME ----#
        global FRAME_TABLE_INVOICE, TREEVIEW_INVOICE, TREEVIEW_SCROLL_INVOICE
        FRAME_TABLE_INVOICE = Frame(FRAME_TAB_INVOICE, bg="white")
        TREEVIEW_INVOICE = ttk.Treeview(FRAME_TABLE_INVOICE, selectmode='browse')
        TREEVIEW_SCROLL_INVOICE = ttk.Scrollbar(FRAME_TABLE_INVOICE, orient="vertical", command=TREEVIEW_INVOICE.yview)
        TREEVIEW_INVOICE.configure(yscrollcommand=TREEVIEW_SCROLL_INVOICE.set)

        TREEVIEW_INVOICE["columns"] = ("one", "two", "three")
        TREEVIEW_INVOICE.column("#0", width=15)
        TREEVIEW_INVOICE.column("one", width=20)
        TREEVIEW_INVOICE.column("two", width=20)
        TREEVIEW_INVOICE.column("three", width=20)

        TREEVIEW_INVOICE.heading("#0", text="Invoice#", anchor="w")
        TREEVIEW_INVOICE.heading("one", text="Customer", anchor="w")
        TREEVIEW_INVOICE.heading("two", text="Cashier", anchor="w")
        TREEVIEW_INVOICE.heading("three", text="Date", anchor="w")

        TREEVIEW_INVOICE.insert('', 'end', text="--", values=["No Data", "Found", "---"])

        TREEVIEW_INVOICE.bind("<Double-1>", invoice_selected_item)

        global FRAME_INVOICE_TAB1
        FRAME_INVOICE_TAB1 = Frame(TAB_INVOICE_INVOICE, bg="black")

        global FRAME_INVOICE_TAB1_WIDGET
        FRAME_INVOICE_TAB1_WIDGET = Frame(FRAME_INVOICE_TAB1, bg="white")

        global FRAME_INVOICE_TAB1_TOP
        FRAME_INVOICE_TAB1_TOP = Frame(FRAME_INVOICE_TAB1, bg="white")

        global ICON_INVOICE, LOGO_INVOICE
        ICON_INVOICE = PhotoImage(file="icon/system_logo.png").subsample(2, 2)
        LOGO_INVOICE = Label(FRAME_INVOICE_TAB1_TOP, image=ICON_INVOICE, bg="white")

        global LABEL_TITLE_INVOICE_TAB1, LABEL_INVOICE_NO, LABEL_INVOICE_DATE
        LABEL_TITLE_INVOICE_TAB1 = Label(FRAME_INVOICE_TAB1_TOP, text="Invoice", font=("algerian", 14), bg="white")

        LABEL_INVOICE_NO = Label(FRAME_INVOICE_TAB1_TOP, text="INVOICE NUMBER", font=("arial", 8), bg="white")

        LABEL_INVOICE_DATE = Label(FRAME_INVOICE_TAB1_TOP, text="DATE OF ISSUE", font=("arial", 8), bg="white")

        global LABEL_INVOICENO_VAL, LABEL_INVOICEDATE_VAL
        LABEL_INVOICENO_VAL = Label(FRAME_INVOICE_TAB1_TOP, text="------", font=("arial", 8), bg="white")

        LABEL_INVOICEDATE_VAL = Label(FRAME_INVOICE_TAB1_TOP, text="------", font=("arial", 8), bg="white")

        global FRAME_INVOICE_INFO_TAB1
        FRAME_INVOICE_INFO_TAB1 = Frame(FRAME_INVOICE_TAB1, bg="white")

        global LABEL_INVOICE_BILL, LABEL_INVOICE_COMPANY
        LABEL_INVOICE_BILL = Label(FRAME_INVOICE_INFO_TAB1, text="BILLED TO", font=("arial", 8), bg="white")

        LABEL_INVOICE_COMPANY = Label(FRAME_INVOICE_INFO_TAB1, text="XWY Corporation", font=("arial", 8), bg="white")

        global LABEL_INVOICE_BILL_NAME_VAL
        LABEL_INVOICE_BILL_NAME_VAL = Label(FRAME_INVOICE_INFO_TAB1, text="-------", font=("arial", 8), bg="white")

        global LABEL_INVOICE_COMPANY_VAL
        LABEL_INVOICE_COMPANY_VAL = Label(FRAME_INVOICE_INFO_TAB1, text="Bacolod City", font=("arial", 8), bg="white")

        global LABEL_INVOICE_BILL_ADDRESS_VAL
        LABEL_INVOICE_BILL_ADDRESS_VAL = Label(FRAME_INVOICE_INFO_TAB1, text="-------", font=("arial", 8), bg="white")

        global LABEL_INVOICE_COMPANY_EMAIL_VAL
        LABEL_INVOICE_COMPANY_EMAIL_VAL = Label(FRAME_INVOICE_INFO_TAB1, text="xyz_corporation@gmail.com",
                                                font=("arial", 8), bg="white")

        # ---- TABLE FRAME ----#
        global FRAME_INVOICE_TREEVIEW_TAB1
        FRAME_INVOICE_TREEVIEW_TAB1 = Frame(FRAME_INVOICE_TAB1, bg="white")

        global INVOICE_TREEVIEW_TAB1
        INVOICE_TREEVIEW_TAB1 = ttk.Treeview(FRAME_INVOICE_TREEVIEW_TAB1, selectmode='browse')

        global TREEVIEW_SCROLL_INVOICE_TAB1
        TREEVIEW_SCROLL_INVOICE_TAB1 = ttk.Scrollbar(FRAME_INVOICE_TREEVIEW_TAB1, orient="vertical",
                                                     command=INVOICE_TREEVIEW_TAB1.yview)
        INVOICE_TREEVIEW_TAB1.configure(yscrollcommand=TREEVIEW_SCROLL_INVOICE_TAB1.set)

        INVOICE_TREEVIEW_TAB1["columns"] = ("one", "two", "three")
        INVOICE_TREEVIEW_TAB1.column("#0", width=100)
        INVOICE_TREEVIEW_TAB1.column("one", width=10)
        INVOICE_TREEVIEW_TAB1.column("two", width=20)
        INVOICE_TREEVIEW_TAB1.column("three", width=20)

        INVOICE_TREEVIEW_TAB1.heading("#0", text="Product name", anchor="w")
        INVOICE_TREEVIEW_TAB1.heading("one", text="Price", anchor="w")
        INVOICE_TREEVIEW_TAB1.heading("two", text="Quantity", anchor="w")
        INVOICE_TREEVIEW_TAB1.heading("three", text="Amount", anchor="w")

        INVOICE_TREEVIEW_TAB1.insert('', 'end', text="", values=["No Data", "---", "---"])

        global FRAME_INVOICE_BOTTOM_TAB1
        FRAME_INVOICE_BOTTOM_TAB1 = Frame(FRAME_INVOICE_TAB1, bg="white")

        global LABEL_INVOICE_TOTAL_TAB1
        LABEL_INVOICE_TOTAL_TAB1 = Label(FRAME_INVOICE_BOTTOM_TAB1, text="INVOICE TOTAL", font=("arial", 8), bg="white")

        global LABEL_INVOICE_TOTAL_TAB1_VAL, LABEL_INVOICE_SUBTOTAL_TAB1, LABEL_INVOICE_SUBTOTAL_TAB1_VAL
        LABEL_INVOICE_TOTAL_TAB1_VAL = Label(FRAME_INVOICE_BOTTOM_TAB1, text="---", font=("arial", 14), bg="white")
        LABEL_INVOICE_SUBTOTAL_TAB1 = Label(FRAME_INVOICE_BOTTOM_TAB1, text="SUBTOTAL", font=("arial", 8), bg="white")
        LABEL_INVOICE_SUBTOTAL_TAB1_VAL = Label(FRAME_INVOICE_BOTTOM_TAB1, text="---", font=("arial", 8), bg="white")

        global LABEL_INVOICE_DISCOUNT_TAB1, LABEL_INVOICE_DISCOUNT_TAB1_VAL
        LABEL_INVOICE_DISCOUNT_TAB1 = Label(FRAME_INVOICE_BOTTOM_TAB1, text="DISCOUNT", font=("arial", 8), bg="white")

        LABEL_INVOICE_DISCOUNT_TAB1_VAL = Label(FRAME_INVOICE_BOTTOM_TAB1, text="---", font=("arial", 8), bg="white")
        LABEL_INVOICE_DISCOUNT_TAB1_VAL.place(x=315, y=15)

        global LABEL_INVOICE_TAX_TAB1, LABEL_INVOICE_TAX_TAB1_VAL
        LABEL_INVOICE_TAX_TAB1 = Label(FRAME_INVOICE_BOTTOM_TAB1, text="TAX RATE", font=("arial", 8), bg="white")
        LABEL_INVOICE_TAX_TAB1_VAL = Label(FRAME_INVOICE_BOTTOM_TAB1, text="---", font=("arial", 8), bg="white")

        global LABEL_INVOICE_ALLTOTAL_TAB1, LABEL_INVOICE_ALLTOTAL_TAB1_VAL
        LABEL_INVOICE_ALLTOTAL_TAB1 = Label(FRAME_INVOICE_BOTTOM_TAB1, text="TOTAL", font=("arial", 8), bg="white")
        LABEL_INVOICE_ALLTOTAL_TAB1_VAL = Label(FRAME_INVOICE_BOTTOM_TAB1, text="---", font=("arial", 8), bg="white")

        # -----------------------------------------------------#
        # --------  PRODUCTS SALES FRAME FORM WIDGET  -------- #
        # -----------------------------------------------------#
        global FORM_SALE_TAB2
        FORM_SALE_TAB2 = Frame(TAB_INVOICE_SALE, bg="white")
        FORM_SALE_TAB2.place(relx=0.01, rely=0.04, relwidth=1, relheight=1)

        global LABEL_SEARCH_FROM_TAB2, LABEL_SEARCH_TO_TAB2
        LABEL_SEARCH_FROM_TAB2 = Label(FORM_SALE_TAB2, text="From:", font=("arial", 11, "bold"), bg="white")
        LABEL_SEARCH_FROM_TAB2.grid(row=0, column=0, sticky="W")

        # ----  PRODUCTS SALES FORM WIDGET ---#
        # search
        global SEARCH_FROM_TAB2_VAR, ENTRY_SEARCH_FROM_TAB2
        SEARCH_FROM_TAB2_VAR = StringVar()
        ENTRY_SEARCH_FROM_TAB2 = DateEntry(FORM_SALE_TAB2, textvariable=SEARCH_FROM_TAB2_VAR, bg="#ddd", font=("arial", 11),
                                       width=15, bd=2, relief=GROOVE, state="readonly", locale='en_US', date_pattern='y-mm-dd')
        ENTRY_SEARCH_FROM_TAB2.grid(row=0, column=1, ipady=3, padx=2, pady=5)


        global LABEL_SEARCH_TO_TAB2
        LABEL_SEARCH_TO_TAB2 = Label(FORM_SALE_TAB2, text="To:", font=("arial", 11, "bold"), bg="white")
        LABEL_SEARCH_TO_TAB2.grid(row=0, column=2, padx=50, sticky="W")

        global SEARCH_TO_TAB2_VAR, ENTRY_SEARCH_TO_TAB2
        SEARCH_TO_TAB2_VAR = StringVar()
        ENTRY_SEARCH_TO_TAB2 = DateEntry(FORM_SALE_TAB2, textvariable=SEARCH_TO_TAB2_VAR, bg="#ddd", font=("arial", 11),
                                     width=20, bd=2, relief=GROOVE, state="readonly", locale='en_US', date_pattern='y-mm-dd')
        ENTRY_SEARCH_TO_TAB2.place(relx=0.37, rely=0.01, relwidth=0.164, relheight=0.068)

        # ---- SEARCH SALES RECORDS ----#
        global ICON_SEARCH_SALES
        ICON_SEARCH_SALES = PhotoImage(file="icon/search.gif")
        BUTTON_SEARCH_SALES = Button(FORM_SALE_TAB2, image=ICON_SEARCH_SALES, bg="#5cb7d8", font=("arial", 11),
                                    width=5, bd=2, relief=GROOVE, command=search_sales)
        BUTTON_SEARCH_SALES.place(relx=0.55, rely=0.01, relwidth=0.05, relheight=0.069)

        # ---- REFRESH SALES RECORDS ----#
        global ICON_REFRESH_SALES
        ICON_REFRESH_SALES = PhotoImage(file="icon/refresh.png")
        BUTTON_REFRESH_SALES = Button(FORM_SALE_TAB2, image=ICON_REFRESH_SALES, bg="#5cb7d8", font=("arial", 11),
                                    width=5, bd=2, relief=GROOVE)
        BUTTON_REFRESH_SALES.place(relx=0.61, rely=0.01, relwidth=0.05, relheight=0.069)
        BUTTON_REFRESH_SALES.bind("<Button-1>", refresh_sales)

        # ---- EXCEL SALES RECORDS ----#
        global ICON_EXCEL_SALES, is_sales_selected
        is_sales_selected = False
        ICON_EXCEL_SALES = PhotoImage(file="icon/excel.png")
        BUTTON_EXCEL_SALES = Button(FORM_SALE_TAB2, image=ICON_EXCEL_SALES, bg="#5cb7d8", font=("arial", 11),
                                     width=5, bd=2, relief=GROOVE)
        BUTTON_EXCEL_SALES.place(relx=0.67, rely=0.01, relwidth=0.05, relheight=0.069)
        BUTTON_EXCEL_SALES.bind("<Button-1>", sales_excel)

        LABEL_SALES_TITLE = Label(FORM_SALE_TAB2, text="SALES REPORT", font=("algerian", 14, "normal"), bg="white")
        LABEL_SALES_TITLE.place(relx=0, rely=0.12)

        # ---- DATE FROM SALES RECORDS ----#
        LABEL_SALES_DATEFROM = Label(FORM_SALE_TAB2, text="From: ", font=("arial", 11, "normal"), bg="white")
        LABEL_SALES_DATEFROM.place(relx=0, rely=0.18)

        global sales_from_var
        sales_from_var = StringVar()
        sales_from_var.set("--------")
        LABEL_SALES_DATEFROM_VAL = Label(FORM_SALE_TAB2, textvariable=sales_from_var, font=("arial", 11, "normal"), bg="white")
        LABEL_SALES_DATEFROM_VAL.place(relx=0.06, rely=0.18)

        # ---- DATE TO SALES RECORDS ----#
        LABEL_SALES_TO = Label(FORM_SALE_TAB2, text="To: ", font=("arial", 11, "normal"), bg="white")
        LABEL_SALES_TO.place(relx=0.16, rely=0.18)

        global sales_to_var
        sales_to_var = StringVar()
        sales_to_var.set("--------")
        LABEL_SALES_TO_VAL = Label(FORM_SALE_TAB2 , textvariable=sales_to_var, font=("arial", 11, "normal"), bg="white")
        LABEL_SALES_TO_VAL.place(relx=0.2, rely=0.18)

        # ---- TABLE FRAME ----#
        global FRAME_TABLE_SALES, TREEVIEW_SALES, TREEVIEW_SCROLL_SALES
        FRAME_TABLE_SALES = Frame(FORM_SALE_TAB2)

        TREEVIEW_SALES = ttk.Treeview(FRAME_TABLE_SALES, selectmode='browse')

        TREEVIEW_SCROLL_SALES = ttk.Scrollbar(FRAME_TABLE_SALES, orient="vertical",
                                                  command=TREEVIEW_SALES.yview)
        TREEVIEW_SALES.configure(yscrollcommand=TREEVIEW_SCROLL_SALES.set)

        TREEVIEW_SALES["columns"] = ("one", "two", "three", "four", "five", "six")
        TREEVIEW_SALES.column("#0", width=50)
        TREEVIEW_SALES.column("one", width=50)
        TREEVIEW_SALES.column("two", width=100)
        TREEVIEW_SALES.column("three", width=180)
        TREEVIEW_SALES.column("four", width=15)
        TREEVIEW_SALES.column("five", width=15)
        TREEVIEW_SALES.column("six", width=10)

        TREEVIEW_SALES.heading("#0", text="Date", anchor="w")
        TREEVIEW_SALES.heading("one", text="Stock no", anchor="w")
        TREEVIEW_SALES.heading("two", text="Product name", anchor="w")
        TREEVIEW_SALES.heading("three", text="Product details", anchor="w")
        TREEVIEW_SALES.heading("four", text="Brand", anchor="w")
        TREEVIEW_SALES.heading("five", text="Price", anchor="w")
        TREEVIEW_SALES.heading("six", text="Quantity", anchor="w")

        TREEVIEW_SALES.insert('', 'end', text="---", values=["---", "Empty", "result", "---", "---", "---"])

        FRAME_TABLE_SALES.place(relx=0, rely=0.25, relwidth=0.98, relheight=0.58)
        TREEVIEW_SCROLL_SALES.pack(side='right', fill='y')
        TREEVIEW_SALES.place(relx=0, rely=0, relwidth=0.975, relheight=1)


        TOTAL_SALES = Label(FORM_SALE_TAB2, text="TOTAL SALES : ", bg="white")
        TOTAL_SALES.place(relx=0.75, rely=0.84, relwidth=0.12, relheight=0.07)

        global total_sales_var
        total_sales_var = StringVar()
        total_sales_label = Label(FORM_SALE_TAB2, textvariable=total_sales_var, bg="white")
        total_sales_label.place(relx=0.86, rely=0.84, relheight=0.07)

        # -----------------------------------------------------#
        # --------  CUSTOMER INFO FRAME FORM WIDGET  -------- #
        # -----------------------------------------------------#
        global FORM_CUSTOMER_TAB3
        FORM_CUSTOMER_TAB3 = Frame(TAB_INVOICE_CUSTOMER, bg="white")

        # ----  CUSTOMER INFO FORM WIDGET ---#
        # search
        global LABEL_SEARCH_CUSTOMER_TAB3
        LABEL_SEARCH_CUSTOMER_TAB3 = Label(FORM_CUSTOMER_TAB3, text="Search:", font=("arial", 11, "bold"),
                                           bg="white")

        global SEARCH_CUSTOMER_TAB3_VAR, ENTRY_SEARCH_CUSTOMER_TAB3
        SEARCH_CUSTOMER_TAB3_VAR = StringVar()
        ENTRY_SEARCH_CUSTOMER_TAB3 = Entry(FORM_CUSTOMER_TAB3, textvariable=SEARCH_CUSTOMER_TAB3_VAR, bg="#ddd",
                                           font=("arial", 11), width=40, bd=2, relief=GROOVE)
        ENTRY_SEARCH_CUSTOMER_TAB3.bind("<Return>", customer_search)

        global ICON_BUTTON_CUSTOMER_TAB3, BUTTON_CUSTOMER_TAB3
        ICON_BUTTON_CUSTOMER_TAB3 = PhotoImage(file="icon/search.gif")
        BUTTON_CUSTOMER_TAB3 = Button(FORM_CUSTOMER_TAB3, image=ICON_BUTTON_CUSTOMER_TAB3, bg="#5cb7d8",
                                      font=("arial", 11), width=5, bd=2, relief=GROOVE)
        BUTTON_CUSTOMER_TAB3.bind("<Button-1>", customer_search)

        global LABEL_CUSTOMER_INFO_NAME, customer_info_name_var
        customer_info_name_var = StringVar()

        LABEL_CUSTOMER_INFO_NAME = Label(FORM_CUSTOMER_TAB3, text="CUSTOMER NAME: ", font=("algerian", 14, "normal"), bg="white")
        LABEL_CUSTOMER_INFO_NAME.place(relx=0, rely=0.14)

        LABEL_CUSTOMER_INFO_NAME_VAL = Label(FORM_CUSTOMER_TAB3, textvariable=customer_info_name_var, font=("arial", 12, "normal"),
                                         bg="white")
        LABEL_CUSTOMER_INFO_NAME_VAL.place(relx=0.21, rely=0.14)

        # ---- TABLE FRAME ----#
        global FRAME_TABLE_CUSTOMER, TREEVIEW_CUSTOMER, TREEVIEW_SCROLL_CUSTOMER
        FRAME_TABLE_CUSTOMER = Frame(FORM_CUSTOMER_TAB3, bg="white")

        TREEVIEW_CUSTOMER = ttk.Treeview(FRAME_TABLE_CUSTOMER, selectmode='browse')

        TREEVIEW_SCROLL_CUSTOMER = ttk.Scrollbar(FRAME_TABLE_CUSTOMER, orient="vertical",
                                              command=TREEVIEW_CUSTOMER.yview)
        TREEVIEW_CUSTOMER.configure(yscrollcommand=TREEVIEW_SCROLL_CUSTOMER.set)

        TREEVIEW_CUSTOMER["columns"] = ("one", "two", "three", "four", "five", "six")
        TREEVIEW_CUSTOMER.column("#0", width=50)
        TREEVIEW_CUSTOMER.column("one", width=50)
        TREEVIEW_CUSTOMER.column("two", width=100)
        TREEVIEW_CUSTOMER.column("three", width=180)
        TREEVIEW_CUSTOMER.column("four", width=15)
        TREEVIEW_CUSTOMER.column("five", width=15)
        TREEVIEW_CUSTOMER.column("six", width=10)

        TREEVIEW_CUSTOMER.heading("#0", text="Date", anchor="w")
        TREEVIEW_CUSTOMER.heading("one", text="Stock no", anchor="w")
        TREEVIEW_CUSTOMER.heading("two", text="Product name", anchor="w")
        TREEVIEW_CUSTOMER.heading("three", text="Product details", anchor="w")
        TREEVIEW_CUSTOMER.heading("four", text="Brand", anchor="w")
        TREEVIEW_CUSTOMER.heading("five", text="Price", anchor="w")
        TREEVIEW_CUSTOMER.heading("six", text="Quantity", anchor="w")

        TREEVIEW_CUSTOMER.insert('', 'end', text="---", values=["---", "Empty", "result", "---", "---", "---"])



        # ----------------------------------------------#
        # --------        DISPLAY WIDGET       -------- #
        # ----------------------------------------------#
        # -----Main frame holds all widget ---#
        # FRAME_PRODUCT_INVOICE.place(relx=0.01, rely=0.1, relwidth=0.98, relheight=1)
        NOTEBOOK_INVOICE_STYLE.configure('TNotebook.Tab', font=('arial', '10', 'bold'))
        NOTEBOOK_INVOICE_STYLE.configure("TNotebook.Tab", padding=[5, 5])
        NOTEBOOK_INVOICE_STYLE.map("TNotebook.Tab", background=[("selected", "blue")])
        NOTEBOOK_INVOICE.place(relx=0, rely=0, relwidth=1, relheight=0.89)
        FORM_INVOICE.place(relx=0.003, rely=0.08, relwidth=0.495, relheight=0.85)
        FORM_INVOICE_WIDGET.place(relx=0.01, rely=0.005, relwidth=0.98, relheight=0.989)
        LABEL_INVOICE_TITLE.place(relx=0, rely=0, relwidth=0.4, relheight=0.1)
        FRAME_INVOICE_TITLE.place(relx=0, rely=0, relwidth=1, relheight=0.1)
        FRAME_TAB_INVOICE.place(relx=0.02, rely=0.1, relwidth=1, relheight=0.9)
        LABEL_MARGIN_INVOICE.grid(row=0, column=0)
        ENTRY_SEARCH_INVOICE.grid(row=1, column=1, ipady=4)
        BUTTON_SEARCH_INVOICE.place(relx=0.8, rely=0.0609, relheight=0.094, relwidth=0.11)
        BUTTON_REFRESH_INVOICE.place(relx=0.76, rely=0, relheight=0.094, relwidth=0.11)
        BUTTON_EXCEL_INVOICE.place(relx=0.88, rely=0, relheight=0.094, relwidth=0.11)
        FRAME_TABLE_INVOICE.place(relx=0, rely=0.2, relwidth=0.96, relheight=0.72)
        TREEVIEW_SCROLL_INVOICE.pack(side='right', fill='y')
        TREEVIEW_INVOICE.place(relx=0, rely=0, relwidth=0.975, relheight=1)
        FRAME_INVOICE_TAB1.place(relx=0.5, rely=0, relwidth=0.5, relheight=1)
        FRAME_INVOICE_TAB1_WIDGET.place(relx=0.005, rely=0.005, relwidth=0.99, relheight=0.99)
        FRAME_INVOICE_TAB1_TOP.place(relx=0.03, rely=0.02, relwidth=0.94)
        LOGO_INVOICE.place(relx=0.9, rely=0, relwidth=0.1)
        LABEL_TITLE_INVOICE_TAB1.grid(row=0, column=0, sticky="w")
        LABEL_INVOICE_NO.grid(row=1, column=0)
        LABEL_INVOICE_DATE.grid(row=1, column=1)
        LABEL_INVOICENO_VAL.grid(row=2, column=0, sticky="w")
        LABEL_INVOICEDATE_VAL.grid(row=2, column=1, sticky="w")
        FRAME_INVOICE_INFO_TAB1.place(relx=0.03, rely=0.2, relwidth=0.94)
        LABEL_INVOICE_BILL.grid(row=0, column=0, sticky="w")
        LABEL_INVOICE_COMPANY.grid(row=0, column=1, padx=160, sticky="w")
        LABEL_INVOICE_BILL_NAME_VAL.grid(row=1, column=0, sticky="w")
        LABEL_INVOICE_COMPANY_VAL.grid(row=1, column=1, padx=160, sticky="w")
        LABEL_INVOICE_BILL_ADDRESS_VAL.grid(row=2, column=0, sticky="w")
        LABEL_INVOICE_COMPANY_EMAIL_VAL.grid(row=2, column=1, padx=160, sticky="w")
        FRAME_INVOICE_TREEVIEW_TAB1.place(relx=0.035, rely=0.36, relwidth=0.94, relheight=0.40)
        TREEVIEW_SCROLL_INVOICE_TAB1.pack(side='right', fill='y')
        INVOICE_TREEVIEW_TAB1.place(relx=0, rely=0, relwidth=0.975, relheight=1)

        FRAME_INVOICE_BOTTOM_TAB1.place(relx=0.03, rely=0.77, relwidth=0.94, relheight=0.2)
        LABEL_INVOICE_TOTAL_TAB1.grid(row=0, column=0, sticky="w")
        LABEL_INVOICE_TOTAL_TAB1_VAL.grid(row=1, column=0, sticky="w")
        LABEL_INVOICE_SUBTOTAL_TAB1.place(x=250, y=0)
        LABEL_INVOICE_SUBTOTAL_TAB1_VAL.place(x=315, y=0)
        LABEL_INVOICE_DISCOUNT_TAB1.place(x=250, y=15)
        LABEL_INVOICE_TAX_TAB1.place(x=250, y=30)
        LABEL_INVOICE_TAX_TAB1_VAL.place(x=315, y=30)
        LABEL_INVOICE_ALLTOTAL_TAB1.place(x=250, y=45)
        LABEL_INVOICE_ALLTOTAL_TAB1_VAL.place(x=315, y=45)

        FORM_CUSTOMER_TAB3.place(relx=0.01, rely=0.04, relwidth=1, relheight=1)
        LABEL_SEARCH_CUSTOMER_TAB3.grid(row=0, column=0, sticky="W")
        ENTRY_SEARCH_CUSTOMER_TAB3.grid(row=0, column=1, ipady=3, padx=4, pady=5)
        BUTTON_CUSTOMER_TAB3.place(relx=0.48, rely=0.0103, relwidth=0.05, relheight=0.065)
        FRAME_TABLE_CUSTOMER.place(relx=0, rely=0.2
                                   , relwidth=0.98, relheight=0.72)
        TREEVIEW_SCROLL_CUSTOMER.pack(side='right', fill='y')
        TREEVIEW_CUSTOMER.place(relx=0, rely=0, relwidth=0.975, relheight=1)

    def display(self):
        FRAME_PRODUCT_INVOICE.place(relx=0.01, rely=0.1, relwidth=0.98, relheight=1)
        LABEL_SUBTITLE.config(text="RECORDS")
        global ICON_RECORDS
        ICON_RECORDS = PhotoImage(file="icon/records.png")
        ICON_PRODUCT_LABEL.config(image=ICON_RECORDS)

    def hide(self):
        FRAME_PRODUCT_INVOICE.place_forget()
        LABEL_SUBTITLE.config(text="")


"""
#====================================================================================================================#
#-----------------------            RECORDS FUNCTIONALITY AREA                                -----------------------#
#====================================================================================================================#
"""
# --------   INVOICE AREA ---#
invoice_array = {}
def invoice_search(event):
    search_val = SEARCH_INVOICE_VAR.get()
    result = controller.get_invoice(search_val)
    table.clear(TREEVIEW_INVOICE)

    if len(result) > 0:
        for row in result:
            invoice_array[row['invoice_no']] = [{
                'id': row['invoice_id'],
                'date': row['invoice_date'],
                'userid': row['user_id'],
                'custid': row['customer_id'],
                'custfname': row['customer_fname'],
                'custmname': row['customer_mname'],
                'custlname': row['customer_lname'],
                'custaddress': row['customer_address'],
                'username': row['user_name']
            }]
            name = "{} {}".format(row['customer_fname'], row['customer_lname'])
            TREEVIEW_INVOICE.insert('', 'end', text=row['invoice_no'],
                                    values=[name, row['user_name'], row['invoice_date']])
    else:
        TREEVIEW_INVOICE.insert('', 'end', text="", values=["No Data", "Found", "---"])

    LABEL_INVOICENO_VAL.config(text="------")
    LABEL_INVOICEDATE_VAL.config(text="------")
    LABEL_INVOICE_BILL_NAME_VAL.config(text="---")
    LABEL_INVOICE_TOTAL_TAB1_VAL.config(text="---")
    LABEL_INVOICE_DISCOUNT_TAB1_VAL.config(text="---")
    LABEL_INVOICE_ALLTOTAL_TAB1_VAL.config(text="---")
    LABEL_INVOICE_TAX_TAB1_VAL.config(text="---")
    LABEL_INVOICE_SUBTOTAL_TAB1_VAL.config(text="---")
    LABEL_INVOICE_TOTAL_TAB1_VAL.config(text="---")

    table.clear(INVOICE_TREEVIEW_TAB1)

def refresh_invoice(event):
    global is_invoice_selected
    table.clear(TREEVIEW_INVOICE)
    TREEVIEW_INVOICE.insert('', 'end', text="--", values=["No Data", "---", "---"])
    is_invoice_selected = False

def invoice_selected_item(event):
    global is_invoice_selected, item_id
    is_invoice_selected = True
    item = TREEVIEW_INVOICE.identify("item", event.x, event.y)
    item_id = TREEVIEW_INVOICE.item(item)['text']

    for row in INVOICE_TREEVIEW_TAB1.get_children():
        INVOICE_TREEVIEW_TAB1.delete(row)

    if item_id == "":
        messagebox.showwarning("Warning", "No Data Found!", parent=FRAME_MAIN)
        return False

    result = controller.get_invoice_sale(item_id)
    LABEL_INVOICEDATE_VAL.config(text=invoice_array[item_id][0]['date'])
    LABEL_INVOICENO_VAL.config(text=item_id)
    LABEL_INVOICE_BILL_NAME_VAL.config(text="{} {}".format(invoice_array[item_id][0]['custfname'], invoice_array[item_id][0]['custlname']))
    LABEL_INVOICE_BILL_ADDRESS_VAL.config(text=invoice_array[item_id][0]['custaddress'])
    total = 0.0

    global excel_record
    excel_record = result


    for row in result:
        amount = float(row['sales_price']) * float(row['sales_quantity'])
        total += amount
        INVOICE_TREEVIEW_TAB1.insert('', 'end', text=row['product_name'],
                                   values=[row['sales_price'], row['sales_quantity'], "{:.2f}".format(amount)])


    LABEL_INVOICE_SUBTOTAL_TAB1_VAL.config(text="{:.2f}".format(total))
    LABEL_INVOICE_TOTAL_TAB1_VAL.config(text="{:.2f}".format(total))
    LABEL_INVOICE_DISCOUNT_TAB1_VAL.config(text="0")
    LABEL_INVOICE_TAX_TAB1_VAL.config(text="0")
    LABEL_INVOICE_ALLTOTAL_TAB1_VAL.config(text="{:.2f}".format(total))


def invoice_excel(event):
    global is_invoice_selected, excel_record

    if is_invoice_selected == True:
        try:
            with OpenKey(HKEY_CURRENT_USER, 'SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders') as key:
                downloads_dir = QueryValueEx(key, '{374DE290-123F-4565-9164-39C4925E467B}')[0]

            wb = openpyxl.Workbook()
            sheet = wb.active

            # ---  ROW 1 --- #
            sheet.merge_cells('A1:C1')
            sheet.cell(row=1, column=1).value = "Invoice report"
            sheet.cell(row=1, column=1).font = Font(size=20)

            # ---  ROW 3 --- #
            sheet.merge_cells('A3:B3')
            sheet.cell(row=3, column=1).value = "INVOICE NUMBER"
            sheet.cell(row=3, column=1).font = Font(size=12, bold=True)

            sheet.merge_cells('C3:D3')
            sheet.cell(row=3, column=3).value = "DATE ISSUE"
            sheet.cell(row=3, column=3).font = Font(size=12, bold=True)

            # ---  ROW 4 --- #
            sheet.merge_cells('A4:B4')
            sheet.cell(row=4, column=1).value = item_id
            sheet.cell(row=4, column=1).font = Font(size=12)

            sheet.merge_cells('C4:D4')
            sheet.cell(row=4, column=3).value = invoice_array[item_id][0]['date']
            sheet.cell(row=4, column=3).font = Font(size=12)

            # ---  ROW 6 --- #
            sheet.merge_cells('A6:B6')
            sheet.cell(row=6, column=1).value = "BILLED TO"
            sheet.cell(row=6, column=1).font = Font(size=12)

            sheet.merge_cells('D6:E6')
            sheet.cell(row=6, column=4).value = "XYZ CORPORATION"
            sheet.cell(row=6, column=4).font = Font(size=12)

            # ---  ROW 7 --- #
            sheet.merge_cells('A7:B7')
            sheet.cell(row=7, column=1).value = "{} {}".format(invoice_array[item_id][0]['custfname'], invoice_array[item_id][0]['custlname'])
            sheet.cell(row=7, column=1).font = Font(size=12)

            sheet.merge_cells('D7:E7')
            sheet.cell(row=7, column=4).value = "Bacolod city"
            sheet.cell(row=7, column=4).font = Font(size=12)

            # ---  ROW 8 --- #
            sheet.merge_cells('A8:C8')
            sheet.cell(row=8, column=1).value = invoice_array[item_id][0]['custaddress']
            sheet.cell(row=8, column=1).font = Font(size=12)

            sheet.merge_cells('D8:F8')
            sheet.cell(row=8, column=4).value = "xyz_corporation@gmail.com"
            sheet.cell(row=8, column=4).font = Font(size=12)

            # ---  ROW 10 --- #
            sheet.merge_cells('A10:C10')
            sheet.cell(row=10, column=1).value = "Product"
            sheet.cell(row=10, column=1).font = Font(size=12, bold=True)

            sheet.cell(row=10, column=4).value = "Price"
            sheet.cell(row=10, column=4).font = Font(size=12, bold=True)

            sheet.cell(row=10, column=5).value = "Quantity"
            sheet.cell(row=10, column=5).font = Font(size=12, bold=True)

            sheet.cell(row=10, column=6).value = "Amount"
            sheet.cell(row=10, column=6).font = Font(size=12, bold=True)
            total = 0
            row_current = 10

            for row in excel_record:
                amount = float(row['sales_price']) * float(row['sales_quantity'])
                total += amount
                row_current += 1

                # ---  ROW 11 --- #
                sheet.merge_cells('A{}:C{}'.format(row_current, row_current))
                sheet.cell(row=row_current, column=1).value = row['product_name']
                sheet.cell(row=row_current, column=1).font = Font(size=12)

                sheet.cell(row=row_current, column=4).value = row['sales_price']
                sheet.cell(row=row_current, column=4).font = Font(size=12)

                sheet.cell(row=row_current, column=5).value = row['sales_quantity']
                sheet.cell(row=row_current, column=5).font = Font(size=12)
                sheet.cell(row=row_current, column=5).alignment = Alignment(horizontal='center')

                sheet.cell(row=row_current, column=6).value = "{:.2f}".format(amount)
                sheet.cell(row=row_current, column=6).font = Font(size=12)


            # ---  ROW 15 --- #
            sheet.merge_cells('A{}:C{}'.format(row_current + 2, row_current + 2))
            sheet.cell(row=(row_current + 2), column=1).value = "INVOICE TOTAL"
            sheet.cell(row=(row_current + 2), column=1).font = Font(size=12)

            sheet.merge_cells('D{}:E{}'.format(row_current + 2, row_current + 2))
            sheet.cell(row=(row_current + 2), column=4).value = "SUBTOTAL"
            sheet.cell(row=(row_current + 2), column=4).font = Font(size=12)

            sheet.cell(row=(row_current + 2), column=6).value = "{:.2f}".format(total)
            sheet.cell(row=(row_current + 2), column=6).font = Font(size=12)

            # ---  ROW 16 --- #
            sheet.merge_cells('A{}:C{}'.format(row_current + 3, row_current + 3))
            sheet.cell(row=(row_current + 3), column=1).value = "{:.2f}".format(total)
            sheet.cell(row=(row_current + 3), column=1).font = Font(size=18)
            sheet.cell(row=(row_current + 3), column=1).alignment = Alignment(horizontal='center')

            sheet.merge_cells('D{}:E{}'.format(row_current + 3, row_current + 3))
            sheet.cell(row=(row_current + 3), column=4).value = "DISCOUNT"
            sheet.cell(row=(row_current + 3), column=4).font = Font(size=12)

            sheet.cell(row=(row_current + 3), column=6).value = "0"
            sheet.cell(row=(row_current + 3), column=6).font = Font(size=12)

            # ---  ROW 17 --- #
            sheet.merge_cells('D{}:E{}'.format(row_current + 4, row_current + 4))
            sheet.cell(row=(row_current + 4), column=4).value = "TAX RATE"
            sheet.cell(row=(row_current + 4), column=4).font = Font(size=12)

            sheet.cell(row=(row_current + 4), column=6).value = "0"
            sheet.cell(row=(row_current + 4), column=6).font = Font(size=12)

            # ---  ROW 18 --- #
            sheet.merge_cells('D{}:E{}'.format(row_current + 5, row_current + 5))
            sheet.cell(row=(row_current + 5), column=4).value = "TOTAL"
            sheet.cell(row=(row_current + 5), column=4).font = Font(size=12)

            sheet.cell(row=(row_current + 5), column=6).value = "{:.2f}".format(total)
            sheet.cell(row=(row_current + 5), column=6).font = Font(size=12)

            wb.save('{}/{}'.format(downloads_dir, "invoice_report.xlsx"))
            messagebox.showinfo("Information message", "Invoice report was succeffully saved in download folder.")
        except:
            messagebox.showerror("Error message", "Invoice report not saved.")
    else:
        messagebox.showwarning("Warning message", "Select invoice no first.")


# --------- SALES ------------#
def refresh_sales(event):
    global is_sales_selected
    table.clear(TREEVIEW_SALES)
    TREEVIEW_SALES.insert('', 'end', text="---", values=["---", "Empty", "result", "---", "---", "---"])
    sales_from_var.set("--------")
    sales_to_var.set("--------")
    total_sales_var.set("")
    is_sales_selected = False


def search_sales():
    global is_sales_selected, excel_sales_record
    is_sales_selected = True
    table.clear(TREEVIEW_SALES)
    date_from = SEARCH_FROM_TAB2_VAR.get()
    date_to = SEARCH_TO_TAB2_VAR.get()
    resp = controller.search_sales_bydate(date_from, date_to)
    sales_from_var.set(SEARCH_FROM_TAB2_VAR.get())
    sales_to_var.set(SEARCH_TO_TAB2_VAR.get())
    total_sales = 0

    if len(resp) > 0:
        excel_sales_record = resp
        for row in resp:
            TREEVIEW_SALES.insert('', 'end', text=row['sales_date'],
                             values=[row['product_id'], row['product_name'],
                                     row['product_details'], row['product_brand'], row['sales_price'], row['sales_quantity']])
            total_sales = total_sales + (float(row['sales_price']) * int(row['sales_quantity']))

        total_sales_var.set("{:.2f}".format(total_sales))
    else:
        TREEVIEW_SALES.insert('', 'end', text="---", values=["---", "Empty", "result", "---", "---", "---"])
        total_sales_var.set("")

def dateformat(d):
    ddd = d.split("/")
    yy = ddd[2]

    dd = "{}{}".format("0", ddd[1]) if int(ddd[1]) < 10 else ddd[1]
    mm = "{}{}".format("0", ddd[0]) if int(ddd[0]) < 10 else ddd[0]

    return "{}-{}-{}".format(yy, mm, dd)

def sales_excel(event):
    global is_sales_selected, excel_sales_record

    if is_sales_selected == True:
        try:
            with OpenKey(HKEY_CURRENT_USER, 'SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders') as key:
                downloads_dir = QueryValueEx(key, '{374DE290-123F-4565-9164-39C4925E467B}')[0]

            wb = openpyxl.Workbook()
            sheet = wb.active

            # ---  ROW 1 --- #
            sheet.merge_cells('A1:C1')
            sheet.cell(row=1, column=1).value = "Sales report"
            sheet.cell(row=1, column=1).font = Font(size=20)

            # ---  ROW 3 --- #
            sheet.cell(row=3, column=1).value = "From: "
            sheet.cell(row=3, column=1).font = Font(size=12, bold=True)

            sheet.cell(row=3, column=2).value = SEARCH_FROM_TAB2_VAR.get()
            sheet.cell(row=3, column=2).font = Font(size=12)

            sheet.cell(row=3, column=4).value = "To: "
            sheet.cell(row=3, column=4).font = Font(size=12, bold=True)

            sheet.cell(row=3, column=5).value = SEARCH_TO_TAB2_VAR.get()
            sheet.cell(row=3, column=5).font = Font(size=12)

            sheet.cell(row=5, column=1).value = "Date"
            sheet.cell(row=5, column=1).font = Font(size=12, bold=True)

            sheet.merge_cells('B{}:C{}'.format(5, 5))
            sheet.cell(row=5, column=2).value = "Stock no"
            sheet.cell(row=5, column=2).font = Font(size=12, bold=True)

            sheet.merge_cells('D{}:E{}'.format(5, 5))
            sheet.cell(row=5, column=4).value = "Product name"
            sheet.cell(row=5, column=4).font = Font(size=12, bold=True)
            #

            sheet.merge_cells('F{}:H{}'.format(5, 5))
            sheet.cell(row=5, column=6).value = "Product details"
            sheet.cell(row=5, column=6).font = Font(size=12, bold=True)

            sheet.cell(row=5, column=9).value = "Brand"
            sheet.cell(row=5, column=9).font = Font(size=12, bold=True)

            sheet.cell(row=5, column=10).value = "Price"
            sheet.cell(row=5, column=10).font = Font(size=12, bold=True)

            sheet.cell(row=5, column=11).value = "Quantity"
            sheet.cell(row=5, column=11).font = Font(size=12, bold=True)


            row_current = 6
            total_sales_excel = 0

            for row in excel_sales_record:
                sheet.cell(row=row_current, column=1).value = row['sales_date']
                sheet.cell(row=row_current, column=1).font = Font(size=12)

                sheet.merge_cells('B{}:C{}'.format(row_current, row_current))
                sheet.cell(row=row_current, column=2).value = row['product_id']
                sheet.cell(row=row_current, column=2).font = Font(size=12)

                sheet.merge_cells('D{}:E{}'.format(row_current, row_current))
                sheet.cell(row=row_current, column=4).value = row['product_name']
                sheet.cell(row=row_current, column=4).font = Font(size=12)

                sheet.merge_cells('F{}:H{}'.format(row_current, row_current))
                sheet.cell(row=row_current, column=6).value = row['product_details']
                sheet.cell(row=row_current, column=6).font = Font(size=12)

                sheet.cell(row=row_current, column=9).value = row['product_brand']
                sheet.cell(row=row_current, column=9).font = Font(size=12)

                sheet.cell(row=row_current, column=10).value = "{:.2f}".format(row['sales_price'])
                sheet.cell(row=row_current, column=10).font = Font(size=12)

                sheet.cell(row=row_current, column=11).value = row['sales_quantity']
                sheet.cell(row=row_current, column=11).font = Font(size=12)
                sheet.cell(row=row_current, column=11).alignment = Alignment(horizontal='center')
                total_sales_excel = total_sales_excel + (float(row['sales_price']) * int(row['sales_quantity']))
                row_current += 1

            sheet.merge_cells('I{}:J{}'.format(row_current + 2, row_current + 2))
            sheet.cell(row=(row_current + 2), column=9).value = "TOTAL SALES: "
            sheet.cell(row=(row_current + 2), column=9).font = Font(size=12)

            sheet.cell(row=(row_current + 2), column=11).value = "{:.2f}".format(total_sales_excel)
            sheet.cell(row=(row_current + 2), column=11).font = Font(size=12)

            wb.save('{}/{}'.format(downloads_dir, "sales_report.xlsx"))
            messagebox.showinfo("Information message", "Sales report was succeffully saved in download folder.")
        except:
            messagebox.showerror("Error message", "Sales report not saved.")
    else:
        messagebox.showwarning("Warning message", "Select date first.")



# --------- CUSTOMER INFO ---------- #
def customer_search(event):
    search_val = SEARCH_CUSTOMER_TAB3_VAR.get()
    table.clear(TREEVIEW_CUSTOMER)

    result = controller.search_customer_invoice(search_val)

    if len(result) > 0:
        customer_info_name_var.set("{} {}".format(result[0]['customer_fname'].capitalize(), result[0]['customer_lname'].capitalize()))
        total = 0.0

        for row in result:
            total += float(row['sales_price'])
            TREEVIEW_CUSTOMER.insert('', 'end', text=row['invoice_date'],
                                      values=["{} / {}".format(row['product_id'], row['invoice_no']),
                                              row['product_name'], row['product_details'],
                                              row['product_brand'], "{:.2f}".format(row['sales_price']),
                                              row['sales_quantity']])
        TREEVIEW_CUSTOMER.insert('', 'end', text="", values=["", "", "", "", "", ""])
        TREEVIEW_CUSTOMER.insert('', 'end', text="", values=["", "", "", "TOTAL", "{:.2f}".format(total), ""])

    else:
        messagebox.showwarning("Warning", "No data found!.", parent=FRAME_BODY_PADDING)
        TREEVIEW_CUSTOMER.insert('', 'end', text="", values=["---", "No data found!.", "---", "---", "---", "---"])

"""
#====================================================================================================================#
#-----------------------            USERS CLASS AREA                                          -----------------------#
#====================================================================================================================#
"""
class User:
    def __init__(self):
        pass

    def show(self):
        # ---- FORM FRAME ----#
        global FRAME_FORM_USER
        FRAME_FORM_USER = Frame(FRAME_BODY_PADDING, bg="white")
        # FRAME_FORM_USER.place(relx=0.01, rely=0.1, relwidth=0.98, relheight=1)

        # ---- TAB FRAME ----#
        global NOTEBOOK_STYLE_USER, NOTEBOOK_USER
        NOTEBOOK_STYLE_USER = ttk.Style()
        NOTEBOOK_STYLE_USER.configure('TNotebook.Tab', font=('arial', '10', 'bold'))
        NOTEBOOK_STYLE_USER.configure("TNotebook.Tab", padding=[5, 5])
        NOTEBOOK_STYLE_USER.map("TNotebook.Tab", background=[("selected", "blue")])
        NOTEBOOK_USER = ttk.Notebook(FRAME_FORM_USER)

        global TAB_USER_REGISTER
        TAB_USER_REGISTER = Frame(NOTEBOOK_USER, bg="white")
        NOTEBOOK_USER.add(TAB_USER_REGISTER, text="User register")

        global TAB_USER_RECORD
        TAB_USER_RECORD = Frame(NOTEBOOK_USER, bg="white")
        NOTEBOOK_USER.add(TAB_USER_RECORD, text="User records")
        NOTEBOOK_USER.place(relx=0, rely=0, relwidth=1, relheight=0.89)

        global FRAME_USER_REGISTER
        FRAME_USER_REGISTER = Frame(TAB_USER_REGISTER, bg="white")
        FRAME_USER_REGISTER.place(relx=0.01, rely=0.04, relwidth=1, relheight=1)

        # -----------------------------------------------------#
        # --------  USER REGISTRATION FORM WIDGET     -------- #
        # -----------------------------------------------------#
        FRAME_USER_REGISTER = Frame(TAB_USER_REGISTER, bg="white")
        FRAME_USER_REGISTER.place(relx=0.1, rely=0.05, relwidth=1, relheight=1)

        # ----  FORM WIDGET ---#
        # user full name
        global LABEL_USER_FULLNAME
        LABEL_USER_FULLNAME = Label(FRAME_USER_REGISTER, text="Fullname:", font=("arial", 11, "bold"), bg="white")
        LABEL_USER_FULLNAME.grid(row=0, column=0, sticky="W")

        global user_fullname_var, ENTRY_USER_FULLNAME
        user_fullname_var = StringVar()
        ENTRY_USER_FULLNAME = Entry(FRAME_USER_REGISTER, textvariable=user_fullname_var, font=("arial", 11), width=60,
                                    bd=2, relief=GROOVE)
        ENTRY_USER_FULLNAME.grid(row=0, column=1, ipady=3, padx=4, pady=5)

        # user position
        global LABEL_USER_POSITION
        LABEL_USER_POSITION = Label(FRAME_USER_REGISTER, text="Position:", font=("arial", 11, "bold"), bg="white")
        LABEL_USER_POSITION.grid(row=1, column=0, sticky="W")

        global user_position_var, COMBOBOX_POSITION
        user_position_var = StringVar()
        position_list = ["Administrator", "Cashier"]
        user_position_var.set(position_list[0])
        COMBOBOX_POSITION = ttk.Combobox(FRAME_USER_REGISTER, textvariable=user_position_var, state="readonly",
                                         font=("arial", 11),
                                         width=58)
        COMBOBOX_POSITION['values'] = position_list
        COMBOBOX_POSITION.grid(row=1, column=1, ipady=3, padx=4, pady=5)

        # user gender
        global LABEL_USER_GENDER
        LABEL_USER_GENDER = Label(FRAME_USER_REGISTER, text="Gender:", font=("arial", 11, "bold"), bg="white")
        LABEL_USER_GENDER.grid(row=2, column=0, sticky="W")

        global user_gender_var, COMBOBOX_USER_GENDER
        user_gender_var = StringVar()
        gender_list = ["Male", "Female"]
        user_gender_var.set(gender_list[0])
        COMBOBOX_USER_GENDER = ttk.Combobox(FRAME_USER_REGISTER, textvariable=user_gender_var, state="readonly",
                                            font=("arial", 11),
                                            width=58)
        COMBOBOX_USER_GENDER['values'] = gender_list
        COMBOBOX_USER_GENDER.grid(row=2, column=1, ipady=3, padx=40, pady=5)

        # user address
        global LABEL_USER_ADDRESS
        LABEL_USER_ADDRESS = Label(FRAME_USER_REGISTER, text="Address:", font=("arial", 11, "bold"), bg="white")
        LABEL_USER_ADDRESS.grid(row=3, column=0, sticky="W")

        global user_address_var, ENTRY_USER_ADDRESS
        user_address_var = StringVar()
        ENTRY_USER_ADDRESS = Entry(FRAME_USER_REGISTER, textvariable=user_address_var, font=("arial", 11), width=60,
                                   bd=2,
                                   relief=GROOVE)
        ENTRY_USER_ADDRESS.grid(row=3, column=1, ipady=3, padx=40, pady=5)

        # user contact no
        global LABEL_USER_CONTACT
        LABEL_USER_CONTACT = Label(FRAME_USER_REGISTER, text="Contact no:", font=("arial", 11, "bold"), bg="white")
        LABEL_USER_CONTACT.grid(row=4, column=0, sticky="W")

        global user_contact_var, ENTRY_USER_CONTACT
        user_contact_var = StringVar()
        ENTRY_USER_CONTACT = Entry(FRAME_USER_REGISTER, textvariable=user_contact_var, font=("arial", 11), width=60,
                                   bd=2,
                                   relief=GROOVE)
        ENTRY_USER_CONTACT.grid(row=4, column=1, ipady=3, padx=40, pady=5)

        # username
        global LABEL_USER_USERAME
        LABEL_USER_USERAME = Label(FRAME_USER_REGISTER, text="Username:", font=("arial", 11, "bold"), bg="white")
        LABEL_USER_USERAME.grid(row=5, column=0, sticky="W")

        global user_username_var, ENTRY_USER_USERAME
        user_username_var = StringVar()
        ENTRY_USER_USERAME = Entry(FRAME_USER_REGISTER, textvariable=user_username_var, font=("arial", 11), width=60,
                                   bd=2,
                                   relief=GROOVE)
        ENTRY_USER_USERAME.grid(row=5, column=1, ipady=3, padx=40, pady=5)

        # password
        global LABEL_USER_PASSWORD
        LABEL_USER_PASSWORD = Label(FRAME_USER_REGISTER, text="Password:", font=("arial", 11, "bold"), bg="white")
        LABEL_USER_PASSWORD.grid(row=6, column=0, sticky="W")

        global user_password_var, ENTRY_USER_PASSWORD
        user_password_var = StringVar()
        ENTRY_USER_PASSWORD = Entry(FRAME_USER_REGISTER, textvariable=user_password_var, font=("arial", 11), width=60,
                                    bd=2,
                                    relief=GROOVE)
        ENTRY_USER_PASSWORD.grid(row=6, column=1, ipady=3, padx=40, pady=5)

        # ---- FRAME CANCEL AND SUBMIT -----#
        global FRAME_USER_BUTTONS
        FRAME_USER_BUTTONS = Frame(FRAME_USER_REGISTER, bg="white")
        FRAME_USER_BUTTONS.place(relx=0.3, rely=0.7)

        # cancel button
        global USER_BUTTON_CANCEL
        USER_BUTTON_CANCEL = Button(FRAME_USER_BUTTONS, text="CANCEL", width=10, bg='red', fg="white",
                                    font=("arial", 11),
                                    relief=GROOVE)
        USER_BUTTON_CANCEL.pack(side=LEFT, pady=15, padx=10, ipady=2)
        USER_BUTTON_CANCEL.bind("<Button-1>", user_cancel_button)

        # submit button
        global USER_BUTTON_SAVE
        USER_BUTTON_SAVE = Button(FRAME_USER_BUTTONS, text="SUBMIT", width=10, bg="blue", fg="white",
                                  font=("arial", 11),
                                  relief=GROOVE)
        USER_BUTTON_SAVE.pack(side=LEFT, padx=10, ipady=2)
        USER_BUTTON_SAVE.bind("<Button-1>", user_submit_button)

        # -----------------------------------------------------#
        # -------  USER RECORDS  FRAME FORM WIDGET    -------- #
        # -----------------------------------------------------#
        FRAME_USER_RECORDS = Frame(TAB_USER_RECORD, bg="white")
        FRAME_USER_RECORDS.place(relx=0.03, rely=0.05, relwidth=0.95, relheight=1)

        # ----  USER RECORDS FORM WIDGET ---#
        # search
        global search_user_var
        search_user_var = StringVar()
        ENTRY_SEARCH_USER = Entry(FRAME_USER_RECORDS, textvariable=search_user_var, bg="#ddd", font=("arial", 11),
                                  width=40,
                                  bd=2,
                                  relief=GROOVE)
        ENTRY_SEARCH_USER.grid(row=0, column=1, ipady=5, padx=4, pady=5)
        ENTRY_SEARCH_USER.bind("<KeyRelease>", user_search)

        BUTTON_SEARCH_USER = Button(FRAME_USER_RECORDS, image=ICON_SEARCH, bg="#5cb7d8", font=("arial", 11), width=5,
                                    bd=2,
                                    relief=GROOVE)
        BUTTON_SEARCH_USER.place(relx=0.4516, rely=0.0102, relwidth=0.05, relheight=0.076)
        BUTTON_SEARCH_USER.bind("<Button-1>", user_search)

        # ---- TABLE FRAME ----#
        FRAME_USER_TABLE = Frame(FRAME_USER_RECORDS)
        FRAME_USER_TABLE.place(relx=0, rely=0.13, relwidth=1, relheight=0.8)

        global TREE_USER_RECORDS
        TREE_USER_RECORDS = ttk.Treeview(FRAME_USER_TABLE, selectmode='browse')
        SCROLLBAR_USER_RECORDS = ttk.Scrollbar(FRAME_USER_TABLE, orient="vertical", command=TREE_USER_RECORDS.yview)
        SCROLLBAR_USER_RECORDS.pack(side='right', fill='y')
        TREE_USER_RECORDS.configure(yscrollcommand=SCROLLBAR_USER_RECORDS.set)

        TREE_USER_RECORDS["columns"] = ("one", "two", "three", "four", "five", "six")
        TREE_USER_RECORDS.column("#0", width=50)
        TREE_USER_RECORDS.column("one", width=20)
        TREE_USER_RECORDS.column("two", width=50)
        TREE_USER_RECORDS.column("three", width=100)
        TREE_USER_RECORDS.column("four", width=50)
        TREE_USER_RECORDS.column("five", width=50)
        TREE_USER_RECORDS.column("six", width=50)

        TREE_USER_RECORDS.heading("#0", text="Fullname", anchor="w")
        TREE_USER_RECORDS.heading("one", text="Gender", anchor="w")
        TREE_USER_RECORDS.heading("two", text="Position", anchor="w")
        TREE_USER_RECORDS.heading("three", text="Address", anchor="w")
        TREE_USER_RECORDS.heading("four", text="Contact#", anchor="w")
        TREE_USER_RECORDS.heading("five", text="Username", anchor="w")
        TREE_USER_RECORDS.heading("six", text="Password", anchor="w")
        TREE_USER_RECORDS.bind("<Double-1>", user_selected_item)

        TREE_USER_RECORDS.place(relx=0, rely=0, relwidth=0.975, relheight=0.96)

    def modal(self):
        # ---- WINDOW TOP CONFIG ----#
        global MODAL_USER
        MODAL_USER = Toplevel(FRAME_MAIN)
        config = window_config(MODAL_USER, 600, 550)

        # ----  MODAL USER TOP --- #
        FRAME_MODAL_USER_TOP = Frame(MODAL_USER, bg="#5cb7d8")
        FRAME_MODAL_USER_TOP.place(relx=0, rely=0, relwidth=1, relheight=0.08)

        ICON_UPDATE_USER = PhotoImage(file="icon/register.png")
        LABEL_ICON_UPDATE_USER = Label(FRAME_MODAL_USER_TOP, image=ICON_UPDATE_USER, bg="#5cb7d8")
        LABEL_ICON_UPDATE_USER.place(relx=0.02, rely=0.01, relwidth=0.08, relheight=1)

        LABEL_TOP_TITLE = Label(FRAME_MODAL_USER_TOP, text="UPDATE USER", font=("algerian", 12), bg="#5cb7d8")
        LABEL_TOP_TITLE.place(relx=0.092, rely=0.5, relwidth=0.18, relheight=1, anchor="w")

        # ---- MODAL USER BODY ---- #
        FRAME_MODAL_USER_BODY = Frame(MODAL_USER, bg="white")
        FRAME_MODAL_USER_BODY.place(relx=0.1, rely=0.14, relwidth=0.8, relheight=0.78)

        global FRAME_USER_UPDATE_FORM
        FRAME_USER_UPDATE_FORM = Frame(FRAME_MODAL_USER_BODY, bg="white")
        FRAME_USER_UPDATE_FORM.place(relx=0.1, rely=0.04, relwidth=1)

        # user full name
        LABEL_USER_UPDATE_FULLNAME = Label(FRAME_USER_UPDATE_FORM, text="Fullname:", font=("arial", 10, "bold"),
                                           bg="white")
        LABEL_USER_UPDATE_FULLNAME.grid(row=0, column=0, sticky="W")

        global user_update_fullname_var
        user_update_fullname_var = StringVar()
        ENTRY_UPDATE_USER_FULLNAME = Entry(FRAME_USER_UPDATE_FORM, textvariable=user_update_fullname_var,
                                           font=("arial", 10), width=37,
                                           bd=2, relief=GROOVE)
        ENTRY_UPDATE_USER_FULLNAME.grid(row=0, column=1, ipady=3, padx=4, pady=5)

        # user position
        LABEL_USER_UPDATE_POSITION = Label(FRAME_USER_UPDATE_FORM, text="Position:", font=("arial", 10, "bold"),
                                           bg="white")
        LABEL_USER_UPDATE_POSITION.grid(row=1, column=0, sticky="W")

        global user_update_position_var
        user_update_position_var = StringVar()
        position_list = ["Administrator", "Cashier"]
        user_update_position_var.set(position_list[0])
        COMBOBOX_USER_UPDATE_POSITION = ttk.Combobox(FRAME_USER_UPDATE_FORM, textvariable=user_update_position_var,
                                                     state="readonly",
                                                     font=("arial", 10),
                                                     width=35)
        COMBOBOX_USER_UPDATE_POSITION['values'] = position_list
        COMBOBOX_USER_UPDATE_POSITION.grid(row=1, column=1, ipady=3, padx=4, pady=5)

        # user gender
        LABEL_USER_UPADTE_GENDER = Label(FRAME_USER_UPDATE_FORM, text="Gender:", font=("arial", 10, "bold"),
                                         bg="white")
        LABEL_USER_UPADTE_GENDER.grid(row=2, column=0, sticky="W")

        global user_update_gender_var, COMBOBOX_USER_UPDATE_GENDER
        user_update_gender_var = StringVar()
        gender_list = ["Male", "Female"]
        user_update_gender_var.set(gender_list[0])
        COMBOBOX_USER_UPDATE_GENDER = ttk.Combobox(FRAME_USER_UPDATE_FORM, textvariable=user_update_gender_var,
                                                   state="readonly",
                                                   font=("arial", 10),
                                                   width=35)
        COMBOBOX_USER_UPDATE_GENDER['values'] = gender_list
        COMBOBOX_USER_UPDATE_GENDER.grid(row=2, column=1, ipady=3, padx=40, pady=5)

        # user address
        LABEL_USER_UPADTE_ADDRESS = Label(FRAME_USER_UPDATE_FORM, text="Address:", font=("arial", 10, "bold"),
                                          bg="white")
        LABEL_USER_UPADTE_ADDRESS.grid(row=3, column=0, sticky="W")

        global user_update_address_var, ENTRY_USER_UPDATE_ADDRESS
        user_update_address_var = StringVar()
        ENTRY_USER_UPDATE_ADDRESS = Entry(FRAME_USER_UPDATE_FORM, textvariable=user_update_address_var,
                                          font=("arial", 10), width=37,
                                          bd=2,
                                          relief=GROOVE)
        ENTRY_USER_UPDATE_ADDRESS.grid(row=3, column=1, ipady=3, padx=40, pady=5)

        # user contact no
        LABEL_USER_UPDATE_CONTACT = Label(FRAME_USER_UPDATE_FORM, text="Contact no:", font=("arial", 10, "bold"),
                                          bg="white")
        LABEL_USER_UPDATE_CONTACT.grid(row=4, column=0, sticky="W")

        global user_update_contact_var, ENTRY_USER_UPDATE_CONTACT
        user_update_contact_var = StringVar()
        ENTRY_USER_UPDATE_CONTACT = Entry(FRAME_USER_UPDATE_FORM, textvariable=user_update_contact_var,
                                          font=("arial", 10), width=37,
                                          bd=2,
                                          relief=GROOVE)
        ENTRY_USER_UPDATE_CONTACT.grid(row=4, column=1, ipady=3, padx=40, pady=5)

        # username
        LABEL_USER_UPADTE_USERAME = Label(FRAME_USER_UPDATE_FORM, text="Username:", font=("arial", 10, "bold"),
                                          bg="white")
        LABEL_USER_UPADTE_USERAME.grid(row=5, column=0, sticky="W")

        global user_update_username_var, ENTRY_USER_UPDATE_USERAME
        user_update_username_var = StringVar()
        ENTRY_USER_UPDATE_USERAME = Entry(FRAME_USER_UPDATE_FORM, textvariable=user_update_username_var,
                                          font=("arial", 10), width=37, bd=2, relief=GROOVE)
        ENTRY_USER_UPDATE_USERAME.grid(row=5, column=1, ipady=3, padx=40, pady=5)

        # password
        LABEL_USER_UPDATE_PASSWORD = Label(FRAME_USER_UPDATE_FORM, text="Password:", font=("arial", 10, "bold"),
                                           bg="white")
        LABEL_USER_UPDATE_PASSWORD.grid(row=6, column=0, sticky="W")

        global user_update_password_var, ENTRY_USER_UPDATE_PASSWORD
        user_update_password_var = StringVar()
        ENTRY_USER_UPDATE_PASSWORD = Entry(FRAME_USER_UPDATE_FORM, textvariable=user_update_password_var,
                                           font=("arial", 10), width=37,
                                           bd=2,
                                           relief=GROOVE)
        ENTRY_USER_UPDATE_PASSWORD.grid(row=6, column=1, ipady=3, padx=40, pady=5)


        global hidden_value_user_id
        hidden_value_user_id = StringVar()
        HIDDEN_USER_ID = Entry(FRAME_USER_UPDATE_FORM, textvariable=hidden_value_user_id)
        # -- status

        global user_update_status_var, COMBOBOX_USER_UPDATE_STATUS
        user_update_status_var = StringVar()
        status_list = ["Active", "Inactive"]
        user_update_status_var.set(status_list[0])
        COMBOBOX_USER_UPDATE_STATUS = ttk.Combobox(FRAME_USER_UPDATE_FORM, textvariable=user_update_status_var,
                                                   state="readonly", font=("arial", 10), width=35)
        COMBOBOX_USER_UPDATE_STATUS['values'] = status_list
        COMBOBOX_USER_UPDATE_STATUS.grid(row=7, column=1, ipady=3, padx=40, pady=5)

        LABEL_USER_UPDATE_STATUS = Label(FRAME_USER_UPDATE_FORM, text="Status:", font=("arial", 10, "bold"),
                                         bg="white")
        LABEL_USER_UPDATE_STATUS.grid(row=7, column=0, sticky="W")

        # ---- FRAME CANCEL AND SUBMIT -----#
        global FRAME_USER_BUTTONS
        FRAME_USER_BUTTONS = Frame(FRAME_MODAL_USER_BODY, bg="white")
        FRAME_USER_BUTTONS.place(relx=0.3, rely=0.8)

        # cancel button
        global USER_BUTTON_CANCEL
        USER_BUTTON_CANCEL = Button(FRAME_USER_BUTTONS, text="CANCEL", width=10, bg='red', fg="white",
                                    font=("arial", 11), relief=GROOVE)
        USER_BUTTON_CANCEL.pack(side=LEFT, pady=15, padx=10, ipady=2)
        USER_BUTTON_CANCEL.bind("<Button-1>", closing_modal_user_key)

        # submit button
        global USER_BUTTON_SAVE
        USER_BUTTON_SAVE = Button(FRAME_USER_BUTTONS, text="SUBMIT", width=10, bg="blue", fg="white",
                                  font=("arial", 11), relief=GROOVE)
        USER_BUTTON_SAVE.pack(side=LEFT, padx=10, ipady=2)
        USER_BUTTON_SAVE.bind("<Button-1>", user_update_submit_button)

        MODAL_USER.grab_set()
        MODAL_USER.bind_all('<Escape>', closing_modal_user_key)
        MODAL_USER.geometry('%dx%d+%d+%d' % (config['WIDTH'], config['HEIGHT'], config['x'], config['y']))
        MODAL_USER.protocol("WM_DELETE_WINDOW", closing_modal_user_window)
        # MODAL_USER.mainloop()

    def display(self):
        FRAME_FORM_USER.place(relx=0.01, rely=0.1, relwidth=0.98, relheight=1)
        LABEL_SUBTITLE.config(text="USER ACCOUNT")
        global ICON_USERS
        ICON_USERS = PhotoImage(file="icon/users.png")
        ICON_PRODUCT_LABEL.config(image=ICON_USERS)

    def hide(self):
        FRAME_FORM_USER.place_forget()
        LABEL_SUBTITLE.config(text="")


"""
#====================================================================================================================#
#-----------------------            USERS FUNCTIONALITY AREA                                  -----------------------#
#====================================================================================================================#
"""
def user_cancel_button(event):
    for frame in FRAME_USER_REGISTER.winfo_children():
        if frame.winfo_class() in ["Entry", "Spinbox"]:
            frame.config(bg="white")

    user_password_var.set("")
    user_username_var.set("")
    user_contact_var.set("")
    user_address_var.set("")
    user_gender_var.set("Male")
    user_position_var.set("Administrator")
    user_fullname_var.set("")
    ENTRY_USER_FULLNAME.focus_set()


def user_submit_button(event):
    upass = user_password_var.get()
    uname = user_username_var.get()
    ucntact = user_contact_var.get()
    uaddress = user_address_var.get()
    ugender = user_gender_var.get()
    uposis = "1" if user_position_var.get() == "Administrator" else "0"
    ufullname = user_fullname_var.get()

    error = validation(FRAME_USER_REGISTER)
    if len(error) > 0:
        messagebox.showwarning("Warnig", error[0], parent=FRAME_MAIN)
    else:
        resp = controller.user_add(ufullname, uposis, ugender, uaddress, ucntact, uname, upass)
        if resp == 1:
            user_cancel_button(event)
            messagebox.showinfo("Success", "User account is successfuly saved.", parent=FRAME_MAIN)
        else:
            messagebox.showwarning("Failed", "Failed to add User!", parent=FRAME_MAIN)

user_list = {}
def user_search(event):
    search_val = search_user_var.get()
    result = controller.search_user(search_val)
    user_list.clear()

    for row in TREE_USER_RECORDS.get_children():
        TREE_USER_RECORDS.delete(row)

    if len(result) > 0:
        for row in result:
            user_list[row['user_username']] = [{
                'id': row['user_id'],
                'name': row['user_name'],
                'position': row['user_position'],
                'gender': row['user_gender'],
                'address': row['user_address'],
                'contact': row['user_contact'],
                'uname': row['user_username'],
                'password': row['user_password'],
                'status': row['user_status']
            }]
            TREE_USER_RECORDS.insert('', 'end', text=row['user_name'], values=[row['user_gender'],
                                        "Administrator" if row['user_position']==1 else "Cashier", row['user_address'], row['user_contact'],
                                        row['user_username'], row['user_password']])
    else:

        TREE_USER_RECORDS.insert('', 'end', text="No", values=["Data", "Found", "---", "---", "---", "---"])

def user_selected_item(event):
    item = TREE_USER_RECORDS.identify("item", event.x, event.y)
    uname = TREE_USER_RECORDS.item(item)['values'][4]

    if uname == "---":
        return False

    frame_user.modal()

    user_info = user_list[uname]

    user_update_fullname_var.set(user_info[0]['name'])
    user_update_position_var.set("Administrator" if user_info[0]['position'] == 1 else "Cashier")
    user_update_gender_var.set(user_info[0]['gender'])
    user_update_address_var.set(user_info[0]['address'])
    user_update_contact_var.set(user_info[0]['contact'])
    user_update_username_var.set(user_info[0]['uname'])
    user_update_password_var.set(user_info[0]['password'])
    user_update_status_var.set("Active" if user_info[0]['status'] == 1 else "Inactive")
    hidden_value_user_id.set(user_info[0]['id'])




def closing_modal_user_key(event):
    MODAL_USER.destroy()

def closing_modal_user_window():
    MODAL_USER.destroy()

def user_update_submit_button(event):

    upass = user_update_password_var.get()
    uname = user_update_username_var.get()
    ucntact = user_update_contact_var.get()
    uaddress = user_update_address_var.get()
    ugender = user_update_gender_var.get()
    uposis = "1" if user_update_position_var.get() == "Administrator" else "0"
    ufullname = user_update_fullname_var.get()
    ustatus = "1" if user_update_status_var.get() == "Active" else "0"
    hide_userid = hidden_value_user_id.get()

    error = validation(FRAME_USER_UPDATE_FORM)
    if len(error) > 0:
        messagebox.showwarning("Warning message", error[0], parent=FRAME_MAIN)
    else:
        resp = controller.update_user_info(upass, uname, ucntact, uaddress, ugender, uposis, ufullname, ustatus, hide_userid)
        if resp == 1:
            closing_modal_user_key(event)
            user_search(event)
            messagebox.showinfo("Success message", "User information updated", parent=FRAME_MAIN)

        else:
            messagebox.showerror("Error message", "Failed to update", parent=FRAME_MAIN)



class Table:
    def __init__(self):
        pass

    def clear(self, tbl):
        for rw in tbl.get_children():
            tbl.delete(rw)

controller = Controller()

frame_main = Main()
frame_login = Login()
frame_home = Home()
frame_product = Product()
frame_inventory = Inventory()
frame_invoice = Invoice()
frame_user = User()
table = Table()

frame_login.show()



